import argparse
import csv
import glob
import json
import os
import re
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import load_workbook

DEFAULT_BC_SLEEP_SECONDS = 0.0
DEFAULT_CALL_FOR_PRICING_LABEL = "CALL 605-858-0545 OR EMAIL SUPPORT@HOTELS4HUMANITY.COM FOR QUOTE"


def env_or_default(*keys, default=""):
    for key in keys:
        value = os.getenv(key)
        if value is not None and str(value).strip() != "":
            return value
    return default


def load_env_file(path):
    loaded = 0
    if not os.path.exists(path):
        return loaded
    with open(path, "r", encoding="utf-8") as f:
        for raw_line in f:
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, val = line.split("=", 1)
            key = key.strip()
            val = val.strip()
            if not key or " " in key:
                continue
            if key in os.environ:
                continue
            if (val.startswith('"') and val.endswith('"')) or (val.startswith("'") and val.endswith("'")):
                val = val[1:-1]
            os.environ[key] = val
            loaded += 1
    return loaded


def bootstrap_env():
    candidates = [
        ".env",
        os.path.join("ahscompany_products", ".env"),
    ]
    loaded_total = 0
    for path in candidates:
        loaded_total += load_env_file(path)
    return loaded_total


def normalize_sku(value):
    if value is None:
        return ""
    cleaned = str(value).strip().upper()
    cleaned = re.sub(r"^KAR[\s\-]+", "", cleaned)
    cleaned = re.sub(r"\s+", "", cleaned)
    return cleaned


def compact_sku(value):
    return re.sub(r"[^A-Z0-9]", "", value or "")


def to_float(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text.replace(",", ""))
    except Exception:
        return None


def to_int(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(text)
    except Exception:
        try:
            return int(float(text))
        except Exception:
            return None


def fmt_number(value):
    if value is None:
        return ""
    number = to_float(value)
    if number is None:
        return str(value).strip()
    if number.is_integer():
        return str(int(number))
    return str(number)


def calc_case_cost(unit_cost, case_pack):
    unit_num = to_float(unit_cost)
    pack_num = to_float(case_pack)
    if unit_num is None or pack_num is None:
        return ""
    return fmt_number(unit_num * pack_num)


def ensure_dir(path):
    Path(path).mkdir(parents=True, exist_ok=True)


def read_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def write_json(path, payload):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=True)


def write_csv(path, rows, fieldnames):
    ensure_dir(Path(path).parent)
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def latest_crawl_json(output_dir):
    pattern = os.path.join(output_dir, "Ahs Company Products *.json")
    files = sorted(glob.glob(pattern), key=os.path.getmtime, reverse=True)
    if not files:
        raise FileNotFoundError(f"No crawl output files found with pattern: {pattern}")
    return files[0]


def parse_option_values(value):
    if value is None:
        return []
    if isinstance(value, list):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return []
        try:
            decoded = json.loads(text)
            return decoded if isinstance(decoded, list) else []
        except Exception:
            return []
    return []


def normalize_header_text(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    if not text:
        return ""
    text = text.replace("\r", " ").replace("\n", " ")
    return re.sub(r"\s+", " ", text)


def header_compact(value):
    return re.sub(r"[^a-z0-9]", "", value or "")


def is_product_id_header(value):
    return header_compact(value) == "productid"


def is_case_price_header(value):
    return header_compact(value).startswith("caseprice")


def is_case_pack_header(value):
    compact = header_compact(value)
    return compact.startswith("casepk") or compact.startswith("casepack")


def is_price_each_header(value):
    compact = header_compact(value)
    # Accept forms like "Price /EA", "Price/EA", "Price / EA", plus labels that
    # include extra context text (example: "Price / EA ... Sold 12/CASE").
    return compact.startswith("priceea")


def parse_case_pack_hint(header_value):
    if not header_value:
        return None
    match = re.search(r"sold\s*([0-9]+(?:\.[0-9]+)?)\s*/\s*case", header_value, flags=re.IGNORECASE)
    return match.group(1) if match else None


def parse_case_pack_from_text(value):
    if value is None:
        return None
    text = normalize_header_text(value)
    if not text:
        return None

    patterns = [
        r"\bsold\s+in\s+(?:cs|case)\s+of\s*([0-9]+(?:\.[0-9]+)?)\s*(?:ea|each)?\b",
        r"\bsold\s*([0-9]+(?:\.[0-9]+)?)\s*/\s*case\b",
        r"\bpacked\s*([0-9]+(?:\.[0-9]+)?)\s*/\s*case\b",
        r"\bpacked\s+([0-9]+(?:\.[0-9]+)?)\s+per\s+case\b",
        r"\b([0-9]+(?:\.[0-9]+)?)\s*(?:ea|each)\s*/\s*case\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return match.group(1)
    return None


def parse_case_pack_from_row_cells(row_values):
    for value in row_values or []:
        parsed = parse_case_pack_from_text(value)
        if parsed is not None:
            return parsed
    return None


def parse_size_hint(header_value):
    if not header_value:
        return ""
    text = normalize_header_text(header_value).replace("\u201d", '"').replace("\u201c", '"')
    match = re.search(
        r"([0-9]+(?:\.[0-9]+)?\s*\"?\s*[x]\s*[0-9]+(?:\.[0-9]+)?\s*\"?(?:\s*[x]\s*[0-9]+(?:\.[0-9]+)?\s*\"?)?)",
        text,
    )
    if not match:
        return ""
    size = match.group(1)
    size = re.sub(r"\s*[x]\s*", " x ", size)
    return re.sub(r"\s+", " ", size).strip()


def is_name_header(value):
    compact = header_compact(value)
    if compact.startswith("namecolor"):
        return True
    return compact == "name"


def is_empty_like(value):
    if value is None:
        return True
    text = str(value).strip().lower()
    return text in {"", "n/a", "na", "call", "call for quote", "quote"}


def first_header_index(headers, predicate):
    for idx, text in enumerate(headers):
        if predicate(text):
            return idx
    return None


def first_valid_price(row, price_indexes):
    if not price_indexes:
        return None, None
    for idx in price_indexes:
        if idx >= len(row):
            continue
        value = row[idx]
        if not is_empty_like(value):
            return value, idx
    return None, None


def build_matrix_price_audit(matrix_rows):
    grouped = defaultdict(list)
    for row in matrix_rows:
        grouped[(row.get("sku_norm", ""), row.get("sheet_name", ""))].append(row)

    out = []
    for (sku_norm, sheet_name), rows in sorted(grouped.items()):
        sku_raw = rows[0].get("sku_raw", "")
        names = []
        sizes = []
        case_packs = []
        prices = []
        for row in rows:
            name = row.get("name_color", "")
            if name and name not in names:
                names.append(name)
            size = row.get("size_hint", "")
            if size and size not in sizes:
                sizes.append(size)
            case_pack = row.get("case_pack", "")
            if case_pack and case_pack not in case_packs:
                case_packs.append(case_pack)
            price_num = to_float(row.get("price_each", ""))
            if price_num is not None:
                prices.append(price_num)

        out.append({
            "sku_norm": sku_norm,
            "sku_raw": sku_raw,
            "sheet_name": sheet_name,
            "matrix_rows": len(rows),
            "name_color_rows": len(names),
            "size_variants": len(sizes),
            "sizes": " | ".join(sizes),
            "case_pack_variants": len(case_packs),
            "case_packs": " | ".join(case_packs),
            "min_price_each": fmt_number(min(prices)) if prices else "",
            "max_price_each": fmt_number(max(prices)) if prices else "",
            "name_color_examples": " || ".join(names[:3]),
        })

    return out


def normalize_for_sku_token(value, max_len=30):
    token = re.sub(r"[^A-Z0-9]+", "", str(value or "").upper())
    if max_len and len(token) > max_len:
        token = token[:max_len]
    return token


def normalize_size_token(size_hint):
    text = str(size_hint or "").upper().replace('"', "")
    text = re.sub(r"\s*[X]\s*", "X", text)
    return normalize_for_sku_token(text, max_len=18)


def normalize_case_pack_token(case_pack):
    pack_int = to_int(case_pack)
    if pack_int is not None and pack_int > 0:
        return str(pack_int)
    return normalize_for_sku_token(case_pack, max_len=8)


def trim_color_fragment(text):
    cleaned = str(text or "").replace("\r", " ").replace("\n", " ")
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" ,;/:-")
    if not cleaned:
        return ""
    cleaned = re.sub(r"^\(([^)]+)\)\s*", "", cleaned).strip()
    cleaned = re.sub(r"\b(standard|packed)\b.*$", "", cleaned, flags=re.IGNORECASE).strip(" ,;/:-")
    return cleaned


def simplify_color_label(label):
    raw = trim_color_fragment(label)
    if not raw:
        return ""
    words = re.findall(r"[A-Za-z]+(?:'[A-Za-z]+)?", raw)
    if not words:
        return raw
    keep_two_prefixes = {"LIGHT", "DARK", "BLUE", "DUSK", "COOL"}
    if len(words) >= 2 and words[-2].upper() in keep_two_prefixes:
        return f"{words[-2]} {words[-1]}"
    return words[-1]


PLAIN_COLOR_PHRASES = [
    "blue moon",
    "light sage",
    "dusk rose",
    "cool gray",
    "cool grey",
]

PLAIN_COLOR_WORDS = {
    "white", "beige", "blue", "green", "gray", "grey", "pink", "brown", "black", "clear",
    "frosty", "cream", "parchment", "jade", "bayberry", "dune", "malt", "pearl", "auburn",
    "mocha", "palm", "buff", "steel", "copen", "cork", "mesa", "pine", "cooper", "taupe",
    "snow", "shadow", "cafe", "doeskin", "fog", "halo", "oatmeal", "scone", "harvest",
    "linen", "iron", "surf", "forest", "smoke", "fern", "blush", "peach", "claret",
    "sage", "sand", "gull", "chestnut", "marsh", "kale", "moss", "heather", "champagne",
    "navy", "redwood", "cypress", "marine", "coco", "wheat",
}


def detect_plain_colors(text):
    cleaned = re.sub(r"\([^)]*\)", " ", str(text or ""))
    cleaned = re.sub(r"\s+", " ", cleaned).strip().lower()
    if not cleaned:
        return []

    found = []
    for phrase in PLAIN_COLOR_PHRASES:
        if phrase in cleaned:
            label = " ".join([word.capitalize() for word in phrase.split()])
            if label not in found:
                found.append(label)
            cleaned = cleaned.replace(phrase, " ")

    for word in re.findall(r"[a-z]+", cleaned):
        if word in PLAIN_COLOR_WORDS:
            label = word.capitalize()
            if label not in found:
                found.append(label)
    return found


def extract_color_entries(name_color):
    text = str(name_color or "").replace("\r", " ").replace("\n", " ")
    text = re.sub(r"\s+", " ", text).strip()
    if not text:
        return []

    entries = []
    code_pattern = re.compile(r"([A-Za-z][A-Za-z0-9/&'(). ]{0,40}?)\s*-\s*(\d{1,3})\b")
    for match in code_pattern.finditer(text):
        raw_label = trim_color_fragment(match.group(1))
        color_code = str(match.group(2)).strip()
        if not raw_label:
            continue
        entries.append({
            "color_label_raw": raw_label,
            "color_label": simplify_color_label(raw_label) or raw_label,
            "color_code": color_code,
        })

    # Some rows include extra plain color values without numeric code (example: ", Beige")
    for segment in re.split(r",|\n", text):
        piece = trim_color_fragment(segment)
        if not piece:
            continue
        if re.search(r"\d", piece):
            continue
        if re.search(r"\b(sold|case|window|grommet|magnet|eco|w/)\b", piece, flags=re.IGNORECASE):
            continue
        entries.append({
            "color_label_raw": piece,
            "color_label": simplify_color_label(piece) or piece,
            "color_code": "",
        })

    if not entries:
        for label in detect_plain_colors(text):
            entries.append({
                "color_label_raw": label,
                "color_label": label,
                "color_code": "",
            })

    if not entries:
        fallback = trim_color_fragment(text)
        if fallback:
            entries.append({
                "color_label_raw": fallback,
                "color_label": simplify_color_label(fallback) or fallback,
                "color_code": "",
            })

    deduped = []
    seen = set()
    for item in entries:
        code = str(item.get("color_code", "")).strip()
        label = str(item.get("color_label", "")).strip()
        key = (code, label.lower())
        if key in seen:
            continue
        seen.add(key)
        deduped.append(item)
    return deduped


def build_option_sku(base_sku_norm, color_code, color_label, size_hint, case_pack):
    base = normalize_for_sku_token(base_sku_norm, max_len=40)
    color_int = to_int(color_code)
    if color_int is not None:
        color_part = f"C{color_int:03d}"
    else:
        color_part = f"CL{normalize_for_sku_token(color_label or 'NA', max_len=16) or 'NA'}"
    size_part = f"S{normalize_size_token(size_hint) or 'NA'}"
    pack_part = f"PK{normalize_case_pack_token(case_pack) or 'NA'}"
    return f"{base}-{color_part}-{size_part}-{pack_part}"


def expand_matrix_rows_to_option_skus(matrix_rows):
    out = []
    seen = set()

    for row in matrix_rows:
        base_sku_norm = row.get("sku_norm", "")
        base_sku_raw = row.get("sku_raw", "")
        size_hint = row.get("size_hint", "")
        case_pack = row.get("case_pack", "")
        price_each = row.get("price_each", "")
        name_color = row.get("name_color", "")

        color_entries = extract_color_entries(name_color)
        if not color_entries:
            color_entries = [{"color_label_raw": "", "color_label": "", "color_code": ""}]

        for color in color_entries:
            color_code = color.get("color_code", "")
            color_label = color.get("color_label", "")
            option_sku = build_option_sku(base_sku_norm, color_code, color_label, size_hint, case_pack)
            option_sku_compact = compact_sku(option_sku)
            color_key = f"C{int(color_code):03d}" if to_int(color_code) is not None else (normalize_for_sku_token(color_label, max_len=16) or "NA")

            key = (option_sku, row.get("sheet_name", ""), row.get("data_row", ""), row.get("price_column_index", ""))
            if key in seen:
                continue
            seen.add(key)

            out.append({
                "option_sku": option_sku,
                "option_sku_compact": option_sku_compact,
                "base_sku_norm": base_sku_norm,
                "base_sku_raw": base_sku_raw,
                "sheet_name": row.get("sheet_name", ""),
                "header_row": row.get("header_row", ""),
                "data_row": row.get("data_row", ""),
                "price_column_index": row.get("price_column_index", ""),
                "size_hint": size_hint,
                "case_pack": case_pack,
                "price_each": price_each,
                "sku_case_cost": calc_case_cost(price_each, case_pack),
                "name_color_source": name_color,
                "color_label": color_label,
                "color_code": color_code,
                "color_key": color_key,
                "derivation_rule": "base_sku + color + size + case_pack",
            })

    return out


def build_option_sku_audit(option_rows):
    grouped = defaultdict(list)
    for row in option_rows:
        grouped[(row.get("base_sku_norm", ""), row.get("sheet_name", ""))].append(row)

    out = []
    for (base_sku_norm, sheet_name), rows in sorted(grouped.items()):
        colors = []
        color_codes = []
        sizes = []
        case_packs = []
        prices = []
        for row in rows:
            color = row.get("color_label", "")
            if color and color not in colors:
                colors.append(color)
            color_code = row.get("color_code", "")
            if color_code and color_code not in color_codes:
                color_codes.append(color_code)
            size = row.get("size_hint", "")
            if size and size not in sizes:
                sizes.append(size)
            pack = row.get("case_pack", "")
            if pack and pack not in case_packs:
                case_packs.append(pack)
            p = to_float(row.get("price_each", ""))
            if p is not None:
                prices.append(p)

        out.append({
            "base_sku_norm": base_sku_norm,
            "sheet_name": sheet_name,
            "option_sku_rows": len(rows),
            "unique_option_skus": len({r.get("option_sku", "") for r in rows if r.get("option_sku", "")}),
            "color_variants": len(colors),
            "color_code_variants": len(color_codes),
            "size_variants": len(sizes),
            "case_pack_variants": len(case_packs),
            "colors": " | ".join(colors[:20]),
            "color_codes": " | ".join(color_codes[:20]),
            "sizes": " | ".join(sizes[:20]),
            "case_packs": " | ".join(case_packs[:20]),
            "min_price_each": fmt_number(min(prices)) if prices else "",
            "max_price_each": fmt_number(max(prices)) if prices else "",
        })

    return out


def normalize_ahs_image_url(value, base_url="https://www.ahscompany.com"):
    text = str(value or "").strip()
    if not text:
        return ""
    if text.startswith("http://") or text.startswith("https://"):
        return text
    if text.startswith("//"):
        return f"https:{text}"
    return f"{base_url.rstrip('/')}/{text.lstrip('/')}"


def parse_image_values(value):
    if value is None:
        return []
    if isinstance(value, list):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return []
        try:
            parsed = json.loads(text)
            if isinstance(parsed, list):
                return parsed
        except Exception:
            pass
        if "," in text:
            return [part.strip() for part in text.split(",") if part.strip()]
        return [text]
    return []


def collect_image_urls(anchor_row, variant_rows=None):
    urls = []
    seen = set()
    rows = [anchor_row] + list(variant_rows or [])
    for row in rows:
        for raw in parse_image_values((row or {}).get("images")):
            url = normalize_ahs_image_url(raw)
            if not url or url in seen:
                continue
            seen.add(url)
            urls.append(url)
    return urls


CATEGORY_TOKEN_STOPWORDS = {
    "THE", "AND", "WITH", "FOR", "FROM", "KARTRI", "HANG2IT", "PER", "CASE", "PRICE", "EACH",
    "SHOWER", "CURTAIN", "POLYESTER", "COTTON", "WINDOW", "SNAP", "AWAY", "LINER", "WHITE",
}


def text_tokens(value):
    text = re.sub(r"[^A-Z0-9]+", " ", str(value or "").upper())
    tokens = set()
    for token in text.split():
        if len(token) <= 1:
            continue
        if token in CATEGORY_TOKEN_STOPWORDS:
            continue
        tokens.add(token)
    return tokens


def normalize_category_ids(value):
    out = []
    for raw in (value or []):
        cid = to_int(raw)
        if cid is not None:
            out.append(cid)
    return out


def build_category_path_lookup(categories):
    by_id = {}
    for row in categories or []:
        cid = to_int(row.get("id"))
        if cid is None:
            continue
        by_id[cid] = row

    memo = {}

    def path_for(cid):
        if cid in memo:
            return memo[cid]
        row = by_id.get(cid)
        if not row:
            memo[cid] = str(cid)
            return memo[cid]
        name = str(row.get("name") or cid).strip()
        pid = to_int(row.get("parent_id"))
        if pid and pid in by_id and pid != cid:
            path = f"{path_for(pid)} > {name}"
        else:
            path = name
        memo[cid] = path
        return path

    return {cid: path_for(cid) for cid in by_id.keys()}


def build_product_category_index(products, brand_id=None):
    target_brand_id = to_int(brand_id)
    out = []
    for p in products or []:
        if target_brand_id is not None and to_int(p.get("brand_id")) != target_brand_id:
            continue
        categories = normalize_category_ids(p.get("categories"))
        if not categories:
            continue
        name = str(p.get("name") or "").strip()
        tokens = text_tokens(name)
        if not tokens:
            continue
        out.append({
            "product_id": str(p.get("id") or "").strip(),
            "name": name,
            "sku": str(p.get("sku") or "").strip(),
            "categories": categories,
            "tokens": tokens,
        })
    return out


def score_category_match(candidate_title, candidate_skus, indexed_product):
    cand_tokens = text_tokens(candidate_title)
    prod_tokens = indexed_product.get("tokens", set())
    if not cand_tokens or not prod_tokens:
        return 0.0
    overlap = cand_tokens & prod_tokens
    if not overlap:
        return 0.0

    union = cand_tokens | prod_tokens
    jaccard = len(overlap) / max(1, len(union))
    score = (len(overlap) * 10.0) + (jaccard * 40.0)

    candidate_prefixes = {normalize_sku(s).split("-")[0] for s in (candidate_skus or []) if normalize_sku(s)}
    product_prefix = normalize_sku(indexed_product.get("sku", "")).split("-")[0]
    if product_prefix and product_prefix in candidate_prefixes:
        score += 25.0

    return round(score, 2)


def suggest_categories_for_candidate(candidate_title, candidate_skus, product_category_index, category_paths):
    best = None
    best_score = 0.0

    for row in product_category_index or []:
        score = score_category_match(candidate_title, candidate_skus, row)
        if score <= 0:
            continue
        if score > best_score:
            best_score = score
            best = row

    if not best:
        return {
            "status": "no_match",
            "score": "",
            "matched_product_id": "",
            "matched_product_name": "",
            "matched_product_sku": "",
            "category_ids": [],
            "category_paths": [],
        }

    # Require a minimum score to auto-assign categories.
    assign_threshold = 22.0
    category_ids = best.get("categories", []) if best_score >= assign_threshold else []
    category_paths_list = [category_paths.get(cid, str(cid)) for cid in category_ids]

    return {
        "status": "matched" if category_ids else "weak_match",
        "score": best_score,
        "matched_product_id": best.get("product_id", ""),
        "matched_product_name": best.get("name", ""),
        "matched_product_sku": best.get("sku", ""),
        "category_ids": category_ids,
        "category_paths": category_paths_list,
    }


def create_candidate_category_rows(candidates):
    out = []
    for row in candidates or []:
        payload = row.get("product_payload", {})
        category_match = row.get("category_match", {})
        category_ids = payload.get("categories", []) or []
        images = payload.get("images", []) or []
        out.append({
            "parent_product_id": row.get("parent_product_id", ""),
            "title": row.get("title", ""),
            "source_url": row.get("source_url", ""),
            "payload_sku": payload.get("sku", ""),
            "missing_sku_count": len(row.get("missing_skus", []) or []),
            "price_sheet_match_count": len(row.get("price_sheet_matched_skus", []) or []),
            "price_sheet_missing_count": len(row.get("price_sheet_missing_skus", []) or []),
            "missing_from_price_sheet": row.get("missing_from_price_sheet", False),
            "call_for_pricing_enabled": bool(payload.get("is_price_hidden")),
            "availability": payload.get("availability", ""),
            "price_hidden_label": payload.get("price_hidden_label", ""),
            "category_match_status": category_match.get("status", ""),
            "category_match_score": category_match.get("score", ""),
            "matched_product_id": category_match.get("matched_product_id", ""),
            "matched_product_sku": category_match.get("matched_product_sku", ""),
            "matched_product_name": category_match.get("matched_product_name", ""),
            "category_ids": "|".join([str(v) for v in category_ids]),
            "category_paths": " || ".join(category_match.get("category_paths", []) or []),
            "image_count": len(images),
            "first_image_url": (images[0] or {}).get("image_url", "") if images else "",
        })
    return out


def load_price_list(price_list_path):
    wb = load_workbook(price_list_path, data_only=True, read_only=True)

    by_norm = {}
    comp_to_norm = {}
    rows = []
    matrix_rows = []
    found_header = False

    for ws in wb.worksheets:
        active_columns = None

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            header_raw = [str(v).strip() if v is not None else "" for v in row]
            headers = [normalize_header_text(v) for v in row]

            pid_idx = first_header_index(headers, is_product_id_header)
            if pid_idx is not None:
                price_each_indexes = [idx for idx, h in enumerate(headers) if is_price_each_header(h)]
                price_each_idx = price_each_indexes[0] if price_each_indexes else None
                case_price_idx = first_header_index(headers, is_case_price_header)
                case_pack_idx = first_header_index(headers, is_case_pack_header)
                name_idx = first_header_index(headers, is_name_header)

                # Some sections use a plain "Price" column with Case Price + Case Pk.
                if price_each_idx is None and case_price_idx is not None and case_pack_idx is not None:
                    fallback_idx = first_header_index(headers, lambda h: h == "price")
                    if fallback_idx is not None:
                        price_each_indexes = [fallback_idx]
                        price_each_idx = fallback_idx

                has_price_context = (
                    price_each_idx is not None
                    or (case_price_idx is not None and case_pack_idx is not None)
                )
                if has_price_context:
                    found_header = True
                    case_pack_hints = {}
                    size_hints = {}
                    if case_pack_idx is None:
                        for price_idx in price_each_indexes:
                            case_pack_hints[price_idx] = parse_case_pack_hint(header_raw[price_idx] if price_idx < len(header_raw) else "")
                            size_hints[price_idx] = parse_size_hint(header_raw[price_idx] if price_idx < len(header_raw) else "")

                    active_columns = {
                        "pid_idx": pid_idx,
                        "price_each_idx": price_each_idx,
                        "price_each_indexes": price_each_indexes,
                        "name_idx": name_idx,
                        "case_pack_idx": case_pack_idx,
                        "case_price_idx": case_price_idx,
                        "case_pack_hints": case_pack_hints,
                        "size_hints": size_hints,
                        "sheet_name": ws.title,
                        "header_row": row_idx,
                    }
                    continue

            if not active_columns:
                continue

            pid_raw = row[active_columns["pid_idx"]] if active_columns["pid_idx"] < len(row) else None
            pid = str(pid_raw).strip() if pid_raw is not None else ""
            if not pid:
                continue

            if is_product_id_header(normalize_header_text(pid)):
                continue

            price_each_idx = active_columns["price_each_idx"]
            price_each_indexes = active_columns["price_each_indexes"]
            case_pack_idx = active_columns["case_pack_idx"]
            case_price_idx = active_columns["case_price_idx"]
            name_idx = active_columns["name_idx"]

            price_each = row[price_each_idx] if price_each_idx is not None and price_each_idx < len(row) else None
            case_pack = row[case_pack_idx] if case_pack_idx is not None and case_pack_idx < len(row) else None
            case_price = row[case_price_idx] if case_price_idx is not None and case_price_idx < len(row) else None
            name_color = row[name_idx] if name_idx is not None and name_idx < len(row) else ""

            if is_empty_like(price_each) and price_each_indexes:
                selected_price, selected_price_idx = first_valid_price(row, price_each_indexes)
                if selected_price is not None:
                    price_each = selected_price
                    price_each_idx = selected_price_idx

            if is_empty_like(case_pack) and price_each_idx is not None:
                hint_pack = active_columns["case_pack_hints"].get(price_each_idx)
                if hint_pack is not None:
                    case_pack = hint_pack
            if is_empty_like(case_pack):
                row_pack = parse_case_pack_from_row_cells(row)
                if row_pack is not None:
                    case_pack = row_pack

            if price_each is None and case_price is not None and to_float(case_pack):
                try:
                    price_each = to_float(case_price) / to_float(case_pack)
                except Exception:
                    price_each = None

            if price_each is None and case_pack is None and case_price is None:
                continue

            norm = normalize_sku(pid)
            if not norm:
                continue

            if len(price_each_indexes) > 1:
                for price_idx in price_each_indexes:
                    if price_idx >= len(row):
                        continue
                    price_value = row[price_idx]
                    if is_empty_like(price_value):
                        continue

                    if case_pack_idx is not None:
                        row_case_pack = case_pack
                    else:
                        row_case_pack = active_columns["case_pack_hints"].get(price_idx)
                        if is_empty_like(row_case_pack):
                            row_case_pack = case_pack

                    matrix_rows.append({
                        "sheet_name": active_columns["sheet_name"],
                        "header_row": active_columns["header_row"],
                        "data_row": row_idx,
                        "sku_raw": pid,
                        "sku_norm": norm,
                        "name_color": str(name_color).strip() if name_color is not None else "",
                        "price_column_index": price_idx,
                        "size_hint": active_columns["size_hints"].get(price_idx, ""),
                        "case_pack": fmt_number(row_case_pack),
                        "price_each": fmt_number(price_value),
                        "case_price_calc": calc_case_cost(price_value, row_case_pack),
                    })

            payload = {
                "sku_raw": pid,
                "sku_norm": norm,
                "sku_compact": compact_sku(norm),
                "price_each": fmt_number(price_each),
                "case_pack": fmt_number(case_pack),
                "case_price_raw": fmt_number(case_price),
                "sku_case_cost": calc_case_cost(price_each, case_pack),
            }
            by_norm[norm] = payload
            if payload["sku_compact"]:
                comp_to_norm[payload["sku_compact"]] = norm
            rows.append(payload)

    if not found_header:
        raise RuntimeError(f"Could not find any Product ID + price headers in price list: {price_list_path}")

    matrix_option_rows = expand_matrix_rows_to_option_skus(matrix_rows)
    matrix_option_audit = build_option_sku_audit(matrix_option_rows)

    return {
        "rows": rows,
        "by_norm": by_norm,
        "comp_to_norm": comp_to_norm,
        "matrix_rows": matrix_rows,
        "matrix_audit": build_matrix_price_audit(matrix_rows),
        "matrix_option_rows": matrix_option_rows,
        "matrix_option_audit": matrix_option_audit,
    }


def build_ahs_index(crawl_rows):
    by_norm = {}
    comp_to_norm = {}
    groups = defaultdict(lambda: {"product": None, "variants": []})

    for row in crawl_rows:
        parent_id = str(row.get("parent_product_id") or row.get("product_id") or "").strip()
        if parent_id:
            if row.get("record_type") == "product":
                groups[parent_id]["product"] = row
            elif row.get("record_type") == "variant":
                groups[parent_id]["variants"].append(row)

        sku_raw = row.get("variant_sku") or row.get("part_number") or ""
        norm = normalize_sku(sku_raw)
        if not norm:
            continue

        existing = by_norm.get(norm)
        current_priority = 1 if row.get("record_type") == "variant" else 0
        existing_priority = 1 if existing and existing.get("record_type") == "variant" else 0
        if existing is None or current_priority > existing_priority:
            by_norm[norm] = {
                "sku_raw": sku_raw,
                "sku_norm": norm,
                "sku_compact": compact_sku(norm),
                "record_type": row.get("record_type", ""),
                "product_id": str(row.get("product_id", "")),
                "parent_product_id": str(row.get("parent_product_id", "")),
                "parent_title": row.get("parent_title", "") or row.get("title", ""),
                "title": row.get("title", ""),
                "variant_sku": row.get("variant_sku", ""),
                "price": fmt_number(row.get("price", "")),
                "ahscompany_price": fmt_number(row.get("ahscompany_price", "")),
                "ahscompany_case_price": fmt_number(row.get("ahscompany_case_price", "")),
                "sku_cost": fmt_number(row.get("sku_cost", "")),
                "sku_case_pack": fmt_number(row.get("sku_case_pack", "")),
                "sku_case_cost": fmt_number(row.get("sku_case_cost", "")),
                "stock_status": row.get("stock_status", ""),
                "url": row.get("url", ""),
                "option_values": parse_option_values(row.get("option_values")),
                "row": row,
            }
            comp = compact_sku(norm)
            if comp:
                comp_to_norm[comp] = norm

    return {
        "by_norm": by_norm,
        "comp_to_norm": comp_to_norm,
        "groups": groups,
    }


class BigCommerceClient:
    def __init__(self, store_hash, access_token, base_url=None, timeout=45, sleep_seconds=DEFAULT_BC_SLEEP_SECONDS):
        self.base_url = (base_url or f"https://api.bigcommerce.com/stores/{store_hash}/v3").rstrip("/")
        self.timeout = timeout
        self.sleep_seconds = max(0.0, float(sleep_seconds or 0))
        self.headers = {
            "X-Auth-Token": access_token,
            "Accept": "application/json",
            "Content-Type": "application/json",
        }

    def _request(self, method, path, params=None, payload=None, retries=6):
        if path.startswith("http://") or path.startswith("https://"):
            url = path
        else:
            url = f"{self.base_url}{path}"

        last_error = None
        for attempt in range(1, retries + 1):
            try:
                resp = requests.request(
                    method=method,
                    url=url,
                    headers=self.headers,
                    params=params,
                    json=payload,
                    timeout=self.timeout,
                )

                if resp.status_code in {429, 500, 502, 503, 504}:
                    wait_s = min(25, attempt * 2)
                    time.sleep(wait_s)
                    continue

                if resp.status_code >= 400:
                    raise RuntimeError(f"HTTP {resp.status_code}: {resp.text[:400]}")

                if not resp.text.strip():
                    return {}
                return resp.json()
            except Exception as e:
                last_error = e
                time.sleep(min(12, attempt * 1.5))
        raise RuntimeError(f"BigCommerce {method} {url} failed: {last_error}")

    def get_paginated(self, path, params=None, limit=250):
        page = 1
        out = []
        base_params = dict(params or {})
        while True:
            q = dict(base_params)
            q["limit"] = limit
            q["page"] = page
            payload = self._request("GET", path, params=q)
            rows = payload.get("data", []) or []
            out.extend(rows)

            pagination = ((payload.get("meta") or {}).get("pagination") or {})
            total_pages = pagination.get("total_pages")
            if total_pages is not None:
                if page >= int(total_pages):
                    break
            else:
                if len(rows) < limit:
                    break
            page += 1
        return out

    def fetch_variants(self):
        return self.get_paginated(
            "/catalog/variants",
            params={"include_fields": "id,product_id,sku,price,sale_price,cost_price"},
            limit=250,
        )

    def fetch_products(self):
        return self.get_paginated(
            "/catalog/products",
            params={"include_fields": "id,name,brand_id,categories,availability,is_price_hidden,price_hidden_label,sku,price,sale_price,cost_price"},
            limit=250,
        )

    def fetch_categories(self):
        return self.get_paginated(
            "/catalog/categories",
            params={"include_fields": "id,parent_id,name"},
            limit=250,
        )

    def update_variant(self, product_id, variant_id, payload):
        result = self._request("PUT", f"/catalog/products/{product_id}/variants/{variant_id}", payload=payload)
        if self.sleep_seconds > 0:
            time.sleep(self.sleep_seconds)
        return result

    def update_product(self, product_id, payload):
        result = self._request("PUT", f"/catalog/products/{product_id}", payload=payload)
        if self.sleep_seconds > 0:
            time.sleep(self.sleep_seconds)
        return result


def build_bc_index(variants, products, brand_id=None):
    by_norm = {}
    comp_to_norm = {}
    duplicates = []
    target_brand_id = to_int(brand_id)
    allowed_product_ids = None

    if target_brand_id is not None:
        allowed_product_ids = set()
        for p in products:
            if to_int(p.get("brand_id")) != target_brand_id:
                continue
            pid = str(p.get("id") or "").strip()
            if pid:
                allowed_product_ids.add(pid)

    def save(row):
        sku_raw = row.get("sku")
        norm = normalize_sku(sku_raw)
        if not norm:
            return
        entry = {
            "sku_raw": sku_raw,
            "sku_norm": norm,
            "sku_compact": compact_sku(norm),
            "bc_type": row.get("bc_type"),
            "product_id": row.get("product_id"),
            "variant_id": row.get("variant_id"),
            "bc_price": fmt_number(row.get("price")),
            "bc_sale_price": fmt_number(row.get("sale_price")),
            "bc_cost_price": fmt_number(row.get("cost_price")),
            "name": row.get("name", ""),
            "brand_id": row.get("brand_id", ""),
            "availability": row.get("availability", ""),
            "is_price_hidden": row.get("is_price_hidden", ""),
            "price_hidden_label": row.get("price_hidden_label", ""),
        }
        if norm in by_norm:
            duplicates.append(entry)
            return
        by_norm[norm] = entry
        comp = entry["sku_compact"]
        if comp:
            comp_to_norm[comp] = norm

    for v in variants:
        if allowed_product_ids is not None:
            pid = str(v.get("product_id") or "").strip()
            if pid not in allowed_product_ids:
                continue
        save({
            "bc_type": "variant",
            "product_id": v.get("product_id"),
            "variant_id": v.get("id"),
            "sku": v.get("sku"),
            "price": v.get("price"),
            "sale_price": v.get("sale_price"),
            "cost_price": v.get("cost_price"),
            "name": "",
        })

    for p in products:
        if allowed_product_ids is not None:
            pid = str(p.get("id") or "").strip()
            if pid not in allowed_product_ids:
                continue
        save({
            "bc_type": "product",
            "product_id": p.get("id"),
            "variant_id": "",
            "sku": p.get("sku"),
            "price": p.get("price"),
            "sale_price": p.get("sale_price"),
            "cost_price": p.get("cost_price"),
            "name": p.get("name", ""),
            "brand_id": p.get("brand_id", ""),
            "availability": p.get("availability", ""),
            "is_price_hidden": p.get("is_price_hidden", ""),
            "price_hidden_label": p.get("price_hidden_label", ""),
        })

    return {
        "by_norm": by_norm,
        "comp_to_norm": comp_to_norm,
        "duplicates": duplicates,
    }


def set_to_rows(skus, primary_map, secondary_map=None):
    out = []
    for sku_norm in sorted(skus):
        row = {"sku_norm": sku_norm}
        if primary_map and sku_norm in primary_map:
            row.update(primary_map[sku_norm])
        if secondary_map and sku_norm in secondary_map:
            for k, v in secondary_map[sku_norm].items():
                row[f"secondary_{k}"] = v
        out.append(row)
    return out


def intersection_rows(skus, ahs_map, price_map=None, bc_map=None):
    rows = []
    for sku_norm in sorted(skus):
        row = {"sku_norm": sku_norm}

        ahs_row = (ahs_map or {}).get(sku_norm, {})
        if ahs_row:
            row.update({
                "ahs_sku_raw": ahs_row.get("sku_raw", ""),
                "ahs_record_type": ahs_row.get("record_type", ""),
                "ahs_product_id": ahs_row.get("product_id", ""),
                "ahs_parent_product_id": ahs_row.get("parent_product_id", ""),
                "ahs_title": ahs_row.get("parent_title", "") or ahs_row.get("title", ""),
                "ahs_variant_sku": ahs_row.get("variant_sku", ""),
                "ahs_price": ahs_row.get("ahscompany_price", ""),
                "ahs_cost": ahs_row.get("sku_cost", ""),
                "ahs_case_pack": ahs_row.get("sku_case_pack", ""),
                "ahs_case_cost": ahs_row.get("sku_case_cost", ""),
            })

        price_row = (price_map or {}).get(sku_norm, {})
        if price_row:
            row.update({
                "price_sheet_sku_raw": price_row.get("sku_raw", ""),
                "price_sheet_price_each": price_row.get("price_each", ""),
                "price_sheet_case_pack": price_row.get("case_pack", ""),
                "price_sheet_case_cost": price_row.get("sku_case_cost", ""),
            })

        bc_row = (bc_map or {}).get(sku_norm, {})
        if bc_row:
            row.update({
                "bc_sku_raw": bc_row.get("sku_raw", ""),
                "bc_type": bc_row.get("bc_type", ""),
                "bc_product_id": bc_row.get("product_id", ""),
                "bc_variant_id": bc_row.get("variant_id", ""),
                "bc_price": bc_row.get("bc_price", ""),
                "bc_sale_price": bc_row.get("bc_sale_price", ""),
                "bc_cost_price": bc_row.get("bc_cost_price", ""),
                "bc_name": bc_row.get("name", ""),
            })

        rows.append(row)
    return rows


def build_update_candidates(
    ahs_map,
    price_map,
    bc_map,
    skus_to_update,
    margin_multiplier=1.48,
    max_change_vs_bc_pct=15.0,
    ahs_below_margin_pct=22.0,
    post_ahs_sale_price=False,
):
    candidates = []
    skipped = []

    for sku_norm in sorted(set(skus_to_update)):
        if sku_norm not in bc_map:
            skipped.append({
                "sku_norm": sku_norm,
                "reason": "missing_in_bc_map",
            })
            continue

        ahs = ahs_map.get(sku_norm, {})
        price_row = price_map.get(sku_norm, {})
        bc = bc_map[sku_norm]

        price_case_pack = to_float(price_row.get("case_pack", ""))
        ahs_case_pack = to_float(ahs.get("sku_case_pack", ""))
        case_qty = price_case_pack if price_case_pack is not None else ahs_case_pack

        price_unit_cost = to_float(price_row.get("price_each", ""))
        ahs_unit_cost = to_float(ahs.get("sku_cost", ""))
        sku_unit_cost = price_unit_cost if price_unit_cost is not None else ahs_unit_cost

        ahs_unit_sales = to_float(ahs.get("ahscompany_price", "")) or to_float(ahs.get("price", ""))
        sku_raw = ahs.get("sku_raw") or price_row.get("sku_raw") or bc.get("sku_raw") or ""

        if case_qty is None or case_qty <= 0:
            skipped.append({
                "sku_norm": sku_norm,
                "sku_raw": sku_raw,
                "reason": "missing_case_qty",
                "sku_case_pack": fmt_number(case_qty),
                "price_sheet_case_pack": price_row.get("case_pack", ""),
                "ahs_case_pack": ahs.get("sku_case_pack", ""),
            })
            continue

        if sku_unit_cost is None or sku_unit_cost <= 0:
            skipped.append({
                "sku_norm": sku_norm,
                "sku_raw": sku_raw,
                "reason": "missing_sku_unit_cost",
                "sku_cost": fmt_number(sku_unit_cost),
                "price_sheet_price_each": price_row.get("price_each", ""),
                "ahs_unit_cost": ahs.get("sku_cost", ""),
            })
            continue

        case_cost = round(sku_unit_cost * case_qty, 2)
        margin_case_price = round(case_cost * float(margin_multiplier), 2)
        ahs_case_sales = round(ahs_unit_sales * case_qty, 2) if (ahs_unit_sales is not None and ahs_unit_sales > 0) else None

        # Requested mapping:
        # - margin_case_price -> BigCommerce price
        # - AHS case sales -> optional BigCommerce sale_price
        case_price = margin_case_price
        price_source = "margin_case_price_to_bc_price"

        bc_price_old_num = to_float(bc.get("bc_price", ""))
        target_vs_bc_pct = ""
        guard_enabled = max_change_vs_bc_pct is not None and float(max_change_vs_bc_pct) > 0
        if bc_price_old_num is not None and bc_price_old_num > 0:
            pct_change = ((case_price - bc_price_old_num) / bc_price_old_num) * 100.0
            target_vs_bc_pct = round(pct_change, 2)
            if guard_enabled and abs(pct_change) > float(max_change_vs_bc_pct):
                skipped.append({
                    "sku_norm": sku_norm,
                    "sku_raw": sku_raw,
                    "reason": f"price_change_exceeds_{max_change_vs_bc_pct}pct_guard",
                    "sku_case_pack": fmt_number(case_qty),
                    "sku_cost": fmt_number(sku_unit_cost),
                    "case_qty": fmt_number(case_qty),
                    "sku_unit_cost": fmt_number(sku_unit_cost),
                    "ahs_unit_sales": fmt_number(ahs_unit_sales),
                    "ahs_case_sales": fmt_number(ahs_case_sales),
                    "margin_case_price": fmt_number(margin_case_price),
                    "candidate_price": fmt_number(case_price),
                    "bc_price_old": bc.get("bc_price", ""),
                    "target_vs_bc_pct": fmt_number(target_vs_bc_pct),
                    "price_source": price_source,
                })
                continue

        payload = {
            "price": case_price,
            "cost_price": case_cost,
        }

        sale_price_status = "not_requested"
        sale_price_min_allowed = ""
        if post_ahs_sale_price:
            if ahs_case_sales is None:
                sale_price_status = "missing_ahs_sales"
            else:
                min_allowed = round(margin_case_price * (1 - float(ahs_below_margin_pct) / 100.0), 2)
                sale_price_min_allowed = fmt_number(min_allowed)
                if ahs_case_sales < min_allowed:
                    sale_price_status = f"blocked_below_margin_guard_{ahs_below_margin_pct}pct"
                else:
                    payload["sale_price"] = ahs_case_sales
                    sale_price_status = "included"

        candidates.append({
            "sku_norm": sku_norm,
            "sku_raw": sku_raw,
            "bc_type": bc.get("bc_type", ""),
            "bc_product_id": bc.get("product_id", ""),
            "bc_variant_id": bc.get("variant_id", ""),
            "case_qty": fmt_number(case_qty),
            "sku_unit_cost": fmt_number(sku_unit_cost),
            "ahs_unit_sales": fmt_number(ahs_unit_sales),
            "ahs_case_sales": fmt_number(ahs_case_sales),
            "margin_case_price": fmt_number(margin_case_price),
            "target_vs_bc_pct": fmt_number(target_vs_bc_pct),
            "price_source": price_source,
            "bc_price_old": bc.get("bc_price", ""),
            "bc_sale_price_old": bc.get("bc_sale_price", ""),
            "bc_cost_old": bc.get("bc_cost_price", ""),
            # Keep requested semantics in output file:
            # - margin_case_price is BC base price target
            # - new_price is case sales target
            "new_price": fmt_number(ahs_case_sales),
            "new_bc_price": fmt_number(payload.get("price")),
            "new_cost_price": fmt_number(payload.get("cost_price")),
            "sale_price_status": sale_price_status,
            "sale_price_min_allowed": sale_price_min_allowed,
            "payload": payload,
        })
    return candidates, skipped


def build_create_candidates(
    groups,
    bc_map,
    price_map=None,
    category_index=None,
    category_paths=None,
    default_brand_id=None,
    call_for_pricing_label=DEFAULT_CALL_FOR_PRICING_LABEL,
):
    bc_skus = set(bc_map.keys())
    price_map = price_map or {}
    out = []
    category_index = category_index or []
    category_paths = category_paths or {}
    brand_id_value = to_int(default_brand_id)

    for parent_id, group in groups.items():
        parent = group.get("product")
        variants = group.get("variants", [])
        if not parent and not variants:
            continue

        group_skus = []
        if variants:
            for v in variants:
                sku = v.get("variant_sku") or v.get("part_number")
                norm = normalize_sku(sku)
                if norm:
                    group_skus.append(norm)
        else:
            sku = (parent or {}).get("variant_sku") or (parent or {}).get("part_number")
            norm = normalize_sku(sku)
            if norm:
                group_skus.append(norm)

        if not group_skus:
            continue
        if any(sku in bc_skus for sku in group_skus):
            continue
        group_price_skus = sorted({sku for sku in group_skus if sku in price_map})
        group_missing_price_skus = sorted({sku for sku in group_skus if sku not in price_map})
        missing_from_price_sheet = len(group_price_skus) == 0

        anchor = parent or variants[0]
        base_name = anchor.get("parent_title") or anchor.get("title") or f"AHS {parent_id}"
        base_price = to_float(anchor.get("ahscompany_price") or anchor.get("price"))
        base_cost = to_float(anchor.get("sku_cost"))
        image_urls = collect_image_urls(anchor, variants)
        category_match = suggest_categories_for_candidate(
            candidate_title=base_name,
            candidate_skus=group_skus,
            product_category_index=category_index,
            category_paths=category_paths,
        )

        candidate = {
            "parent_product_id": parent_id,
            "title": base_name,
            "source_url": anchor.get("url", ""),
            "missing_skus": group_skus,
            "price_sheet_matched_skus": group_price_skus,
            "price_sheet_missing_skus": group_missing_price_skus,
            "missing_from_price_sheet": missing_from_price_sheet,
            "category_match": category_match,
            "product_payload": {
                "name": base_name,
                "type": "physical",
                "sku": (anchor.get("variant_sku") or anchor.get("part_number") or "").strip(),
                "price": base_price if base_price is not None else 0,
                "cost_price": base_cost if base_cost is not None else 0,
                "description": anchor.get("description", ""),
            },
        }
        if brand_id_value is not None:
            candidate["product_payload"]["brand_id"] = brand_id_value
        if image_urls:
            candidate["product_payload"]["images"] = [{"image_url": u} for u in image_urls[:20]]
        if category_match.get("category_ids"):
            candidate["product_payload"]["categories"] = category_match["category_ids"]
        if missing_from_price_sheet:
            candidate["product_payload"]["availability"] = "disabled"
            candidate["product_payload"]["is_price_hidden"] = True
            candidate["product_payload"]["price_hidden_label"] = call_for_pricing_label
            candidate["product_payload"]["price"] = 0
            candidate["product_payload"]["cost_price"] = 0

        if variants:
            option_map = defaultdict(list)
            variant_payloads = []
            for v in variants:
                sku = (v.get("variant_sku") or v.get("part_number") or "").strip()
                if not sku:
                    continue

                option_values = parse_option_values(v.get("option_values"))
                clean_ov = []
                for ov in option_values:
                    if not isinstance(ov, dict):
                        continue
                    name = str(ov.get("option_display_name") or "").strip()
                    label = str(ov.get("label") or "").strip()
                    if not name or not label:
                        continue
                    clean_ov.append({"option_display_name": name, "label": label})
                    if label not in option_map[name]:
                        option_map[name].append(label)

                variant_payload = {
                    "sku": sku,
                    "price": to_float(v.get("ahscompany_price") or v.get("price")) or 0,
                    "cost_price": to_float(v.get("sku_cost")) or 0,
                    "option_values": clean_ov,
                }
                variant_payloads.append(variant_payload)

            product_options = []
            for name, labels in option_map.items():
                product_options.append({
                    "display_name": name,
                    "type": "dropdown",
                    "option_values": [{"label": label} for label in labels],
                })

            if product_options:
                candidate["product_payload"]["variants"] = variant_payloads
                candidate["product_payload"]["options"] = product_options

        out.append(candidate)

    return out


def build_disable_missing_price_sheet_candidates(
    bc_map,
    price_map,
    sku_norms_to_disable,
    call_for_pricing_label=DEFAULT_CALL_FOR_PRICING_LABEL,
):
    disable_rows = []
    skipped_placeholder_products = []
    restore_placeholder_products = []
    price_map = price_map or {}
    by_product = defaultdict(lambda: {"product": None, "variants": []})

    for entry in (bc_map or {}).values():
        pid = str(entry.get("product_id") or "").strip()
        if not pid:
            continue
        if entry.get("bc_type") == "product":
            by_product[pid]["product"] = entry
        elif entry.get("bc_type") == "variant":
            by_product[pid]["variants"].append(entry)

    for sku_norm in sorted(set(sku_norms_to_disable or [])):
        bc = (bc_map or {}).get(sku_norm, {})
        if not bc:
            continue

        bc_type = bc.get("bc_type", "")
        product_id = str(bc.get("product_id") or "").strip()
        group = by_product.get(product_id, {"product": None, "variants": []})
        variant_rows = group.get("variants", [])
        variant_matched = sorted({v.get("sku_norm", "") for v in variant_rows if v.get("sku_norm", "") in price_map})
        variant_missing = sorted({v.get("sku_norm", "") for v in variant_rows if v.get("sku_norm", "") and v.get("sku_norm", "") not in price_map})
        has_priced_variant = len(variant_matched) > 0

        # Product-level placeholder SKUs should be ignored when priced variants exist.
        if bc_type == "product" and variant_rows and has_priced_variant:
            skipped_placeholder_products.append({
                "sku_norm": sku_norm,
                "sku_raw": bc.get("sku_raw", ""),
                "bc_type": bc_type,
                "bc_product_id": product_id,
                "reason": "product_placeholder_sku_ignored_due_to_priced_variants",
                "priced_variant_count": len(variant_matched),
                "missing_variant_count": len(variant_missing),
                "priced_variants": "|".join(variant_matched[:30]),
                "missing_variants": "|".join(variant_missing[:30]),
            })

            # If it was previously hidden/disabled by call-for-pricing rule, restore it.
            availability = str(bc.get("availability", "")).strip().lower()
            is_price_hidden = str(bc.get("is_price_hidden", "")).strip().lower() in {"true", "1"}
            label = str(bc.get("price_hidden_label", "")).strip()
            if availability == "disabled" and is_price_hidden and label == call_for_pricing_label:
                restore_placeholder_products.append({
                    "sku_norm": sku_norm,
                    "sku_raw": bc.get("sku_raw", ""),
                    "bc_type": bc_type,
                    "bc_product_id": product_id,
                    "bc_variant_id": "",
                    "reason": "restore_placeholder_product_with_priced_variants",
                    "call_for_pricing_label": call_for_pricing_label,
                    "payload": {
                        "availability": "available",
                        "is_price_hidden": False,
                        "price_hidden_label": "",
                    },
                })
            continue

        if bc_type == "variant":
            payload = {
                # Variant-level purchase block for SKUs missing from price sheet.
                "purchasing_disabled": True,
            }
        else:
            payload = {
                "availability": "disabled",
                "is_price_hidden": True,
                "price_hidden_label": call_for_pricing_label,
            }

        disable_rows.append({
            "sku_norm": sku_norm,
            "sku_raw": bc.get("sku_raw", ""),
            "bc_type": bc_type,
            "bc_product_id": bc.get("product_id", ""),
            "bc_variant_id": bc.get("variant_id", ""),
            "bc_price_old": bc.get("bc_price", ""),
            "bc_sale_price_old": bc.get("bc_sale_price", ""),
            "bc_cost_price_old": bc.get("bc_cost_price", ""),
            "reason": "bc_sku_missing_from_price_sheet",
            "call_for_pricing_label": call_for_pricing_label,
            "payload": payload,
        })
    return disable_rows, skipped_placeholder_products, restore_placeholder_products


def apply_full_payload_updates(client, candidates):
    results = []
    for row in candidates:
        payload = row.get("payload") or {}
        if not payload:
            results.append({**row, "applied": False, "status": "skipped", "error": "empty_payload"})
            continue

        try:
            if row.get("bc_type") == "variant":
                client.update_variant(row.get("bc_product_id"), row.get("bc_variant_id"), payload)
            else:
                client.update_product(row.get("bc_product_id"), payload)
            results.append({**row, "applied": True, "status": "ok", "error": ""})
        except Exception as e:
            results.append({**row, "applied": True, "status": "error", "error": str(e)})
    return results


def apply_updates(client, candidates, update_price=True, update_cost=True):
    results = []
    for row in candidates:
        payload = {}
        if update_price and "price" in row["payload"]:
            payload["price"] = row["payload"]["price"]
        if update_price and "sale_price" in row["payload"]:
            payload["sale_price"] = row["payload"]["sale_price"]
        if update_cost and "cost_price" in row["payload"]:
            payload["cost_price"] = row["payload"]["cost_price"]
        if not payload:
            results.append({**row, "applied": False, "status": "skipped", "error": "empty_payload"})
            continue

        try:
            if row["bc_type"] == "variant":
                client.update_variant(row["bc_product_id"], row["bc_variant_id"], payload)
            else:
                client.update_product(row["bc_product_id"], payload)
            results.append({**row, "applied": True, "status": "ok", "error": ""})
        except Exception as e:
            results.append({**row, "applied": True, "status": "error", "error": str(e)})
    return results


def main():
    bootstrap_env()

    parser = argparse.ArgumentParser(
        description="Reconcile AHS crawl + Kartri price list + BigCommerce SKUs, and optionally apply BC price/cost updates."
    )
    parser.add_argument("--crawl-json", default="", help="Path to AHS crawl JSON. Defaults to latest file in ./output.")
    parser.add_argument("--output-dir", default="output", help="Directory where crawl JSON files are stored.")
    parser.add_argument("--price-list", default="2026 Kartri Price List.xlsx", help="Path to Kartri price list XLSX.")
    parser.add_argument("--reports-dir", default="", help="Output directory for reconciliation reports.")
    parser.add_argument("--store-hash", default=env_or_default("BIGCOMMERCE_STORE_HASH", default=""))
    parser.add_argument("--access-token", default=env_or_default("BCPRODUCT_ACCESS_TOKEN", "BIGCOMMERCE_ACCESS_TOKEN", default=""))
    parser.add_argument("--api-base-url", default=os.getenv("BIGCOMMERCE_API_BASE_URL", ""))
    parser.add_argument(
        "--call-for-pricing-label",
        default=DEFAULT_CALL_FOR_PRICING_LABEL,
        help="Label used when product is hidden/disabled because no price-sheet SKU matched.",
    )
    parser.add_argument(
        "--disable-missing-price-sheet-skus",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Auto-disable existing BigCommerce SKUs not found in price sheet and set call-for-pricing where supported.",
    )
    parser.add_argument(
        "--bc-brand-id",
        type=int,
        default=2945,
        help="Only include BigCommerce products/variants under this brand_id for matching (set 0 to disable filter).",
    )
    parser.add_argument(
        "--bc-sleep-seconds",
        type=float,
        default=float(env_or_default("BC_SLEEP_SECONDS", default=str(DEFAULT_BC_SLEEP_SECONDS))),
        help="Sleep between BigCommerce write calls (seconds).",
    )
    parser.add_argument("--apply-updates", action="store_true", help="Apply BC price/cost updates for matched SKUs.")
    parser.add_argument("--no-update-price", action="store_true", help="Do not update BC price.")
    parser.add_argument("--no-update-cost", action="store_true", help="Do not update BC cost_price.")
    parser.add_argument(
        "--margin-multiplier",
        type=float,
        default=1.48,
        help="Default markup multiplier used when AHS unit sales price is missing.",
    )
    parser.add_argument(
        "--max-change-vs-bc-pct",
        type=float,
        default=15.0,
        help="Safety guard: skip updates where target price change vs current BC price exceeds this percent.",
    )
    parser.add_argument(
        "--ahs-below-margin-pct",
        type=float,
        default=22.0,
        help="When --post-ahs-sale-price is enabled, block sale_price if AHS case sales is more than this percent below margin_case_price.",
    )
    parser.add_argument(
        "--post-ahs-sale-price",
        action="store_true",
        help="Optional: also post AHS case sales to BigCommerce sale_price when it passes --ahs-below-margin-pct guard.",
    )
    parser.add_argument(
        "--update-scope",
        choices=["all_three", "price_bc"],
        default="all_three",
        help="SKU scope for BC updates: all_three (AHS + price + BC) or price_bc (price + BC only).",
    )
    args = parser.parse_args()

    crawl_json = args.crawl_json or latest_crawl_json(args.output_dir)
    if not os.path.exists(crawl_json):
        raise FileNotFoundError(f"Crawl JSON not found: {crawl_json}")
    if not os.path.exists(args.price_list):
        raise FileNotFoundError(f"Price list not found: {args.price_list}")

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    reports_dir = args.reports_dir or os.path.join(args.output_dir, f"kartri_reconcile_{stamp}")
    ensure_dir(reports_dir)

    crawl_rows = read_json(crawl_json)
    ahs = build_ahs_index(crawl_rows)
    price = load_price_list(args.price_list)
    matrix_rows = price.get("matrix_rows", [])
    matrix_audit = price.get("matrix_audit", [])
    matrix_option_rows = price.get("matrix_option_rows", [])
    matrix_option_audit = price.get("matrix_option_audit", [])

    bc = {"by_norm": {}, "comp_to_norm": {}, "duplicates": []}
    disable_candidates = []
    disable_results = []
    placeholder_disable_skipped = []
    restore_placeholder_candidates = []
    restore_placeholder_results = []
    products = []
    categories = []
    bc_category_paths = {}
    bc_product_category_index = []
    bc_enabled = bool(args.store_hash and args.access_token)
    bc_brand_filter = args.bc_brand_id if args.bc_brand_id and int(args.bc_brand_id) > 0 else None
    if bc_enabled:
        client = BigCommerceClient(
            store_hash=args.store_hash,
            access_token=args.access_token,
            base_url=args.api_base_url or None,
            sleep_seconds=args.bc_sleep_seconds,
        )
        print("[INFO] Fetching BigCommerce products...")
        products = client.fetch_products()
        print("[INFO] Fetching BigCommerce categories...")
        categories = client.fetch_categories()
        print("[INFO] Fetching BigCommerce variants...")
        variants = client.fetch_variants()
        bc = build_bc_index(variants, products, brand_id=bc_brand_filter)
        bc_category_paths = build_category_path_lookup(categories)
        bc_product_category_index = build_product_category_index(products, brand_id=bc_brand_filter)
        if bc_brand_filter is not None:
            print(f"[INFO] Loaded BC SKU index (brand_id={bc_brand_filter}): {len(bc['by_norm'])} unique SKUs")
        else:
            print(f"[INFO] Loaded BC SKU index (all brands): {len(bc['by_norm'])} unique SKUs")
        print(f"[INFO] Category matching corpus: {len(bc_product_category_index)} products with categories")
    else:
        print("[WARN] BIGCOMMERCE_STORE_HASH or BCPRODUCT_ACCESS_TOKEN missing; BC comparisons skipped.")

    ahs_skus = set(ahs["by_norm"].keys())
    price_skus = set(price["by_norm"].keys())
    bc_skus = set(bc["by_norm"].keys())

    ahs_not_price = ahs_skus - price_skus
    price_not_ahs = price_skus - ahs_skus
    matched_ahs_price = ahs_skus & price_skus
    ahs_not_bc = ahs_skus - bc_skus if bc_enabled else set()
    price_not_bc = price_skus - bc_skus if bc_enabled else set()
    bc_not_ahs = bc_skus - ahs_skus if bc_enabled else set()
    bc_not_price = bc_skus - price_skus if bc_enabled else set()
    ahs_missing_price_and_bc = ahs_skus - price_skus - bc_skus if bc_enabled else set()
    matched_ahs_bc = ahs_skus & bc_skus if bc_enabled else set()
    matched_price_bc = price_skus & bc_skus if bc_enabled else set()
    matched_all_three = ahs_skus & price_skus & bc_skus if bc_enabled else set()

    write_csv(
        os.path.join(reports_dir, "matched_ahs_price_skus.csv"),
        intersection_rows(matched_ahs_price, ahs["by_norm"], price["by_norm"]),
        fieldnames=[
            "sku_norm", "ahs_sku_raw", "ahs_record_type", "ahs_product_id", "ahs_parent_product_id",
            "ahs_title", "ahs_variant_sku", "ahs_price", "ahs_cost", "ahs_case_pack", "ahs_case_cost",
            "price_sheet_sku_raw", "price_sheet_price_each", "price_sheet_case_pack", "price_sheet_case_cost",
        ],
    )

    write_csv(
        os.path.join(reports_dir, "ahs_skus_not_in_price_list.csv"),
        set_to_rows(ahs_not_price, ahs["by_norm"]),
        fieldnames=[
            "sku_norm", "sku_raw", "record_type", "product_id", "parent_product_id",
            "parent_title", "variant_sku", "sku_cost", "sku_case_pack", "sku_case_cost",
            "ahscompany_price", "ahscompany_case_price", "stock_status", "url",
        ],
    )
    write_csv(
        os.path.join(reports_dir, "price_list_skus_not_in_ahs.csv"),
        set_to_rows(price_not_ahs, price["by_norm"]),
        fieldnames=["sku_norm", "sku_raw", "price_each", "case_pack", "case_price_raw", "sku_case_cost"],
    )

    if matrix_rows:
        write_csv(
            os.path.join(reports_dir, "price_list_matrix_variations.csv"),
            matrix_rows,
            fieldnames=[
                "sheet_name", "header_row", "data_row",
                "sku_raw", "sku_norm", "name_color",
                "price_column_index", "size_hint", "case_pack", "price_each", "case_price_calc",
            ],
        )
        write_json(os.path.join(reports_dir, "price_list_matrix_variations.json"), matrix_rows)

    if matrix_audit:
        write_csv(
            os.path.join(reports_dir, "price_list_matrix_audit.csv"),
            matrix_audit,
            fieldnames=[
                "sku_norm", "sku_raw", "sheet_name",
                "matrix_rows", "name_color_rows", "size_variants", "sizes",
                "case_pack_variants", "case_packs", "min_price_each", "max_price_each",
                "name_color_examples",
            ],
        )
        write_json(os.path.join(reports_dir, "price_list_matrix_audit.json"), matrix_audit)

    if matrix_option_rows:
        write_csv(
            os.path.join(reports_dir, "price_list_option_skus.csv"),
            matrix_option_rows,
            fieldnames=[
                "option_sku", "option_sku_compact",
                "base_sku_norm", "base_sku_raw",
                "sheet_name", "header_row", "data_row", "price_column_index",
                "size_hint", "case_pack", "price_each", "sku_case_cost",
                "name_color_source", "color_label", "color_code", "color_key",
                "derivation_rule",
            ],
        )
        write_json(os.path.join(reports_dir, "price_list_option_skus.json"), matrix_option_rows)

    if matrix_option_audit:
        write_csv(
            os.path.join(reports_dir, "price_list_option_sku_audit.csv"),
            matrix_option_audit,
            fieldnames=[
                "base_sku_norm", "sheet_name", "option_sku_rows", "unique_option_skus",
                "color_variants", "color_code_variants", "size_variants", "case_pack_variants",
                "colors", "color_codes", "sizes", "case_packs",
                "min_price_each", "max_price_each",
            ],
        )
        write_json(os.path.join(reports_dir, "price_list_option_sku_audit.json"), matrix_option_audit)

    if bc_enabled:
        write_csv(
            os.path.join(reports_dir, "ahs_skus_not_in_bigcommerce.csv"),
            set_to_rows(ahs_not_bc, ahs["by_norm"]),
            fieldnames=[
                "sku_norm", "sku_raw", "record_type", "product_id", "parent_product_id",
                "parent_title", "variant_sku", "sku_cost", "sku_case_pack", "sku_case_cost",
                "ahscompany_price", "ahscompany_case_price", "stock_status", "url",
            ],
        )
        write_csv(
            os.path.join(reports_dir, "price_list_skus_not_in_bigcommerce.csv"),
            set_to_rows(price_not_bc, price["by_norm"]),
            fieldnames=["sku_norm", "sku_raw", "price_each", "case_pack", "case_price_raw", "sku_case_cost"],
        )
        write_csv(
            os.path.join(reports_dir, "ahs_missing_from_price_and_bigcommerce.csv"),
            set_to_rows(ahs_missing_price_and_bc, ahs["by_norm"]),
            fieldnames=[
                "sku_norm", "sku_raw", "record_type", "product_id", "parent_product_id",
                "parent_title", "variant_sku", "sku_cost", "sku_case_pack", "sku_case_cost",
                "ahscompany_price", "ahscompany_case_price", "stock_status", "url",
            ],
        )
        write_csv(
            os.path.join(reports_dir, "bigcommerce_skus_not_in_ahs.csv"),
            set_to_rows(bc_not_ahs, bc["by_norm"]),
            fieldnames=["sku_norm", "sku_raw", "bc_type", "product_id", "variant_id", "bc_price", "bc_cost_price", "name"],
        )
        write_csv(
            os.path.join(reports_dir, "bigcommerce_skus_not_in_price_list.csv"),
            set_to_rows(bc_not_price, bc["by_norm"]),
            fieldnames=["sku_norm", "sku_raw", "bc_type", "product_id", "variant_id", "bc_price", "bc_cost_price", "name"],
        )
        if args.disable_missing_price_sheet_skus:
            disable_candidates, placeholder_disable_skipped, restore_placeholder_candidates = build_disable_missing_price_sheet_candidates(
                bc_map=bc["by_norm"],
                price_map=price["by_norm"],
                sku_norms_to_disable=bc_not_price,
                call_for_pricing_label=args.call_for_pricing_label,
            )
            write_csv(
                os.path.join(reports_dir, "bigcommerce_disable_missing_price_sheet_candidates.csv"),
                disable_candidates,
                fieldnames=[
                    "sku_norm", "sku_raw", "bc_type", "bc_product_id", "bc_variant_id",
                    "bc_price_old", "bc_sale_price_old", "bc_cost_price_old",
                    "reason", "call_for_pricing_label",
                ],
            )
            write_json(
                os.path.join(reports_dir, "bigcommerce_disable_missing_price_sheet_candidates.json"),
                disable_candidates,
            )
            if placeholder_disable_skipped:
                write_csv(
                    os.path.join(reports_dir, "bigcommerce_disable_missing_price_sheet_skipped_placeholder_products.csv"),
                    placeholder_disable_skipped,
                    fieldnames=[
                        "sku_norm", "sku_raw", "bc_type", "bc_product_id",
                        "reason", "priced_variant_count", "missing_variant_count",
                        "priced_variants", "missing_variants",
                    ],
                )
                write_json(
                    os.path.join(reports_dir, "bigcommerce_disable_missing_price_sheet_skipped_placeholder_products.json"),
                    placeholder_disable_skipped,
                )
            if restore_placeholder_candidates:
                write_csv(
                    os.path.join(reports_dir, "bigcommerce_restore_placeholder_products_candidates.csv"),
                    restore_placeholder_candidates,
                    fieldnames=[
                        "sku_norm", "sku_raw", "bc_type", "bc_product_id", "bc_variant_id",
                        "reason", "call_for_pricing_label",
                    ],
                )
                write_json(
                    os.path.join(reports_dir, "bigcommerce_restore_placeholder_products_candidates.json"),
                    restore_placeholder_candidates,
                )

        write_csv(
            os.path.join(reports_dir, "matched_ahs_bigcommerce_skus.csv"),
            intersection_rows(matched_ahs_bc, ahs["by_norm"], bc_map=bc["by_norm"]),
            fieldnames=[
                "sku_norm", "ahs_sku_raw", "ahs_record_type", "ahs_product_id", "ahs_parent_product_id",
                "ahs_title", "ahs_variant_sku", "ahs_price", "ahs_cost", "ahs_case_pack", "ahs_case_cost",
                "bc_sku_raw", "bc_type", "bc_product_id", "bc_variant_id", "bc_price", "bc_cost_price", "bc_name",
            ],
        )
        write_csv(
            os.path.join(reports_dir, "matched_price_bigcommerce_skus.csv"),
            intersection_rows(matched_price_bc, {}, price["by_norm"], bc["by_norm"]),
            fieldnames=[
                "sku_norm", "price_sheet_sku_raw", "price_sheet_price_each", "price_sheet_case_pack", "price_sheet_case_cost",
                "bc_sku_raw", "bc_type", "bc_product_id", "bc_variant_id", "bc_price", "bc_cost_price", "bc_name",
            ],
        )
        write_csv(
            os.path.join(reports_dir, "matched_ahs_price_bigcommerce_skus.csv"),
            intersection_rows(matched_all_three, ahs["by_norm"], price["by_norm"], bc["by_norm"]),
            fieldnames=[
                "sku_norm", "ahs_sku_raw", "ahs_record_type", "ahs_product_id", "ahs_parent_product_id",
                "ahs_title", "ahs_variant_sku", "ahs_price", "ahs_cost", "ahs_case_pack", "ahs_case_cost",
                "price_sheet_sku_raw", "price_sheet_price_each", "price_sheet_case_pack", "price_sheet_case_cost",
                "bc_sku_raw", "bc_type", "bc_product_id", "bc_variant_id", "bc_price", "bc_cost_price", "bc_name",
            ],
        )

        if matrix_rows:
            matrix_rows_in_bc = []
            for row in matrix_rows:
                bc_row = bc["by_norm"].get(row.get("sku_norm", ""), {})
                if not bc_row:
                    continue
                matrix_rows_in_bc.append({
                    "sku_norm": row.get("sku_norm", ""),
                    "sku_raw": row.get("sku_raw", ""),
                    "sheet_name": row.get("sheet_name", ""),
                    "name_color": row.get("name_color", ""),
                    "size_hint": row.get("size_hint", ""),
                    "case_pack": row.get("case_pack", ""),
                    "price_each": row.get("price_each", ""),
                    "case_price_calc": row.get("case_price_calc", ""),
                    "bc_type": bc_row.get("bc_type", ""),
                    "bc_product_id": bc_row.get("product_id", ""),
                    "bc_variant_id": bc_row.get("variant_id", ""),
                    "bc_sku_raw": bc_row.get("sku_raw", ""),
                    "bc_price": bc_row.get("bc_price", ""),
                    "bc_sale_price": bc_row.get("bc_sale_price", ""),
                    "bc_cost_price": bc_row.get("bc_cost_price", ""),
                })
            write_csv(
                os.path.join(reports_dir, "price_list_matrix_rows_in_bigcommerce.csv"),
                matrix_rows_in_bc,
                fieldnames=[
                    "sku_norm", "sku_raw", "sheet_name", "name_color", "size_hint",
                    "case_pack", "price_each", "case_price_calc",
                    "bc_type", "bc_product_id", "bc_variant_id", "bc_sku_raw",
                    "bc_price", "bc_sale_price", "bc_cost_price",
                ],
            )

        if matrix_option_rows:
            option_rows_in_bc = []
            for row in matrix_option_rows:
                bc_row = bc["by_norm"].get(row.get("base_sku_norm", ""), {})
                if not bc_row:
                    continue
                option_rows_in_bc.append({
                    "option_sku": row.get("option_sku", ""),
                    "base_sku_norm": row.get("base_sku_norm", ""),
                    "base_sku_raw": row.get("base_sku_raw", ""),
                    "color_label": row.get("color_label", ""),
                    "color_code": row.get("color_code", ""),
                    "size_hint": row.get("size_hint", ""),
                    "case_pack": row.get("case_pack", ""),
                    "price_each": row.get("price_each", ""),
                    "sku_case_cost": row.get("sku_case_cost", ""),
                    "bc_type": bc_row.get("bc_type", ""),
                    "bc_product_id": bc_row.get("product_id", ""),
                    "bc_variant_id": bc_row.get("variant_id", ""),
                    "bc_sku_raw": bc_row.get("sku_raw", ""),
                    "bc_price": bc_row.get("bc_price", ""),
                    "bc_sale_price": bc_row.get("bc_sale_price", ""),
                    "bc_cost_price": bc_row.get("bc_cost_price", ""),
                })
            write_csv(
                os.path.join(reports_dir, "price_list_option_skus_in_bigcommerce.csv"),
                option_rows_in_bc,
                fieldnames=[
                    "option_sku", "base_sku_norm", "base_sku_raw",
                    "color_label", "color_code", "size_hint", "case_pack", "price_each", "sku_case_cost",
                    "bc_type", "bc_product_id", "bc_variant_id", "bc_sku_raw",
                    "bc_price", "bc_sale_price", "bc_cost_price",
                ],
            )

        if args.update_scope == "price_bc":
            skus_to_update = matched_price_bc
        else:
            skus_to_update = matched_all_three
        print(f"[INFO] Update scope: {args.update_scope} ({len(skus_to_update)} SKUs before guards)")

        effective_max_change_guard = None if args.update_scope == "price_bc" else args.max_change_vs_bc_pct
        update_candidates, update_skipped = build_update_candidates(
            ahs_map=ahs["by_norm"],
            price_map=price["by_norm"],
            bc_map=bc["by_norm"],
            skus_to_update=skus_to_update,
            margin_multiplier=args.margin_multiplier,
            max_change_vs_bc_pct=effective_max_change_guard,
            ahs_below_margin_pct=args.ahs_below_margin_pct,
            post_ahs_sale_price=args.post_ahs_sale_price,
        )
        write_csv(
            os.path.join(reports_dir, "bigcommerce_update_candidates.csv"),
            update_candidates,
            fieldnames=[
                "sku_norm", "sku_raw", "bc_type", "bc_product_id", "bc_variant_id",
                "case_qty", "sku_unit_cost", "ahs_unit_sales", "ahs_case_sales", "margin_case_price",
                "target_vs_bc_pct", "price_source", "sale_price_status", "sale_price_min_allowed",
                "bc_price_old", "bc_sale_price_old", "bc_cost_old", "new_price", "new_bc_price", "new_cost_price",
            ],
        )
        write_csv(
            os.path.join(reports_dir, "bigcommerce_update_skipped.csv"),
            update_skipped,
            fieldnames=[
                "sku_norm", "sku_raw", "reason", "sku_case_pack", "sku_cost", "case_qty",
                "sku_unit_cost", "ahs_unit_sales", "ahs_case_sales", "margin_case_price", "candidate_price",
                "bc_price_old", "target_vs_bc_pct", "price_source",
            ],
        )

        if args.apply_updates:
            results = apply_updates(
                client=client,
                candidates=update_candidates,
                update_price=not args.no_update_price,
                update_cost=not args.no_update_cost,
            )
            write_csv(
                os.path.join(reports_dir, "bigcommerce_update_results.csv"),
                results,
                fieldnames=[
                    "sku_norm", "sku_raw", "bc_type", "bc_product_id", "bc_variant_id",
                    "case_qty", "sku_unit_cost", "ahs_unit_sales", "ahs_case_sales", "margin_case_price",
                    "target_vs_bc_pct", "price_source", "sale_price_status", "sale_price_min_allowed",
                    "bc_price_old", "bc_sale_price_old", "bc_cost_old", "new_price", "new_bc_price", "new_cost_price",
                    "applied", "status", "error",
                ],
            )

            if args.disable_missing_price_sheet_skus and disable_candidates:
                disable_results = apply_full_payload_updates(client=client, candidates=disable_candidates)
                write_csv(
                    os.path.join(reports_dir, "bigcommerce_disable_missing_price_sheet_results.csv"),
                    disable_results,
                    fieldnames=[
                        "sku_norm", "sku_raw", "bc_type", "bc_product_id", "bc_variant_id",
                        "bc_price_old", "bc_sale_price_old", "bc_cost_price_old",
                        "reason", "call_for_pricing_label",
                        "applied", "status", "error",
                    ],
                )
            if args.disable_missing_price_sheet_skus and restore_placeholder_candidates:
                restore_placeholder_results = apply_full_payload_updates(client=client, candidates=restore_placeholder_candidates)
                write_csv(
                    os.path.join(reports_dir, "bigcommerce_restore_placeholder_products_results.csv"),
                    restore_placeholder_results,
                    fieldnames=[
                        "sku_norm", "sku_raw", "bc_type", "bc_product_id", "bc_variant_id",
                        "reason", "call_for_pricing_label",
                        "applied", "status", "error",
                    ],
                )
        else:
            print("[INFO] Dry-run mode: no BigCommerce updates applied.")

        create_candidates = build_create_candidates(
            ahs["groups"],
            bc["by_norm"],
            price_map=price["by_norm"],
            category_index=bc_product_category_index,
            category_paths=bc_category_paths,
            default_brand_id=bc_brand_filter,
            call_for_pricing_label=args.call_for_pricing_label,
        )
        create_category_rows = create_candidate_category_rows(create_candidates)
        write_csv(
            os.path.join(reports_dir, "bigcommerce_create_category_suggestions.csv"),
            create_category_rows,
            fieldnames=[
                "parent_product_id", "title", "source_url", "payload_sku", "missing_sku_count",
                "price_sheet_match_count", "price_sheet_missing_count", "missing_from_price_sheet",
                "call_for_pricing_enabled", "availability", "price_hidden_label",
                "category_match_status", "category_match_score",
                "matched_product_id", "matched_product_sku", "matched_product_name",
                "category_ids", "category_paths", "image_count", "first_image_url",
            ],
        )
        write_json(os.path.join(reports_dir, "bigcommerce_create_category_suggestions.json"), create_category_rows)
        write_json(os.path.join(reports_dir, "bigcommerce_create_product_candidates.json"), create_candidates)
    else:
        create_candidates = build_create_candidates(
            ahs["groups"],
            {},
            price_map=price["by_norm"],
            default_brand_id=bc_brand_filter,
            call_for_pricing_label=args.call_for_pricing_label,
        )
        create_category_rows = create_candidate_category_rows(create_candidates)
        write_csv(
            os.path.join(reports_dir, "bigcommerce_create_category_suggestions.csv"),
            create_category_rows,
            fieldnames=[
                "parent_product_id", "title", "source_url", "payload_sku", "missing_sku_count",
                "price_sheet_match_count", "price_sheet_missing_count", "missing_from_price_sheet",
                "call_for_pricing_enabled", "availability", "price_hidden_label",
                "category_match_status", "category_match_score",
                "matched_product_id", "matched_product_sku", "matched_product_name",
                "category_ids", "category_paths", "image_count", "first_image_url",
            ],
        )
        write_json(os.path.join(reports_dir, "bigcommerce_create_category_suggestions.json"), create_category_rows)
        write_json(os.path.join(reports_dir, "bigcommerce_create_product_candidates.json"), create_candidates)

    summary = {
        "crawl_json": crawl_json,
        "price_list": os.path.abspath(args.price_list),
        "reports_dir": os.path.abspath(reports_dir),
        "counts": {
            "ahs_unique_skus": len(ahs_skus),
            "price_list_unique_skus": len(price_skus),
            "price_list_matrix_rows": len(matrix_rows),
            "price_list_matrix_skus": len({row.get("sku_norm", "") for row in matrix_rows if row.get("sku_norm", "")}),
            "price_list_matrix_audit_groups": len(matrix_audit),
            "price_list_option_sku_rows": len(matrix_option_rows),
            "price_list_option_skus": len({row.get("option_sku", "") for row in matrix_option_rows if row.get("option_sku", "")}),
            "price_list_option_skus_with_color_code": len({row.get("option_sku", "") for row in matrix_option_rows if row.get("option_sku", "") and row.get("color_code", "")}),
            "price_list_option_sku_audit_groups": len(matrix_option_audit),
            "bc_unique_skus": len(bc_skus),
            "matched_ahs_price": len(matched_ahs_price),
            "matched_ahs_bigcommerce": len(matched_ahs_bc),
            "matched_price_bigcommerce": len(matched_price_bc),
            "matched_all_three": len(matched_all_three),
            "update_candidates_case_based": len(update_candidates) if bc_enabled else 0,
            "update_skipped_case_based": len(update_skipped) if bc_enabled else 0,
            "create_candidates": len(create_candidates),
            "create_candidates_with_categories": len([r for r in create_candidates if (r.get("product_payload", {}).get("categories"))]),
            "create_candidates_with_images": len([r for r in create_candidates if (r.get("product_payload", {}).get("images"))]),
            "create_candidates_with_category_match": len([r for r in create_candidates if r.get("category_match", {}).get("status") == "matched"]),
            "create_candidates_missing_price_sheet": len([r for r in create_candidates if r.get("missing_from_price_sheet")]),
            "create_candidates_call_for_pricing": len([r for r in create_candidates if r.get("product_payload", {}).get("is_price_hidden")]),
            "bc_disable_missing_price_sheet_candidates": len(disable_candidates) if bc_enabled else 0,
            "bc_disable_missing_price_sheet_results": len(disable_results) if bc_enabled else 0,
            "bc_disable_missing_price_sheet_skipped_placeholder_products": len(placeholder_disable_skipped) if bc_enabled else 0,
            "bc_restore_placeholder_products_candidates": len(restore_placeholder_candidates) if bc_enabled else 0,
            "bc_restore_placeholder_products_results": len(restore_placeholder_results) if bc_enabled else 0,
            "ahs_not_in_price_list": len(ahs_not_price),
            "price_list_not_in_ahs": len(price_not_ahs),
            "ahs_not_in_bigcommerce": len(ahs_not_bc),
            "price_list_not_in_bigcommerce": len(price_not_bc),
            "ahs_missing_from_price_and_bigcommerce": len(ahs_missing_price_and_bc),
            "bigcommerce_not_in_ahs": len(bc_not_ahs),
            "bigcommerce_not_in_price_list": len(bc_not_price),
        },
        "bc_enabled": bc_enabled,
        "bc_duplicates": len(bc.get("duplicates", [])),
        "bc_brand_id_filter": bc_brand_filter,
        "call_for_pricing_label": args.call_for_pricing_label,
        "disable_missing_price_sheet_skus": bool(args.disable_missing_price_sheet_skus),
        "update_scope": args.update_scope,
    }
    write_json(os.path.join(reports_dir, "summary.json"), summary)

    print("\n[SUMMARY]")
    for k, v in summary["counts"].items():
        print(f"- {k}: {v}")
    print(f"\n[OK] Reports written to: {os.path.abspath(reports_dir)}")


if __name__ == "__main__":
    main()
