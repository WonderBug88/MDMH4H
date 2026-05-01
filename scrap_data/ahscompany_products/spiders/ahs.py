import copy
import os
import re
from collections import OrderedDict
from datetime import datetime
from urllib.parse import urljoin

from scrapy import Spider, Request, signals


class AshSpider(Spider):
    name = "ahs"
    base_url = "https://www.ahscompany.com/"
    start_urls = ["https://www.ahscompany.com/Kartri_c_5205.html"]

    custom_settings = {
        # Conservative crawl profile to reduce 429s from AHS.
        'CONCURRENT_REQUESTS': 1,
        'CONCURRENT_REQUESTS_PER_DOMAIN': 1,
        'DOWNLOAD_DELAY': 1.5,
        'RANDOMIZE_DOWNLOAD_DELAY': True,
        'AUTOTHROTTLE_ENABLED': True,
        'AUTOTHROTTLE_START_DELAY': 1.0,
        'AUTOTHROTTLE_MAX_DELAY': 20.0,
        'AUTOTHROTTLE_TARGET_CONCURRENCY': 1.0,
        'RETRY_TIMES': 8,
        'RETRY_HTTP_CODES': [500, 502, 503, 504, 400, 403, 404, 408, 429],

        'FEEDS': {
            f'output/Ahs Company Products {datetime.now().strftime("%d%m%Y%H%M%S")}.json': {
                'format': 'json',
                'fields': ['record_type', 'product_id', 'parent_product_id', 'parent_title',
                           'title', 'variant_sku', 'option_values', 'price', 'retail_price',
                           'brand', 'stock_status', 'part_number', 'sku_cost', 'sku_case_cost', 'sku_case_pack',
                           'ahscompany_price', 'ahscompany_case_price', 'options',
                           'images', 'description', 'category',
                           'sub_category', 'size', 'type', 'material', 'url'],
            }
        }
    }

    headers = {
        'authority': 'www.ahscompany.com',
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'accept-language': 'en-PK,en;q=0.9,ur-PK;q=0.8,ur;q=0.7,en-US;q=0.6',
        'cache-control': 'no-cache',
        'pragma': 'no-cache',
        'upgrade-insecure-requests': '1',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
    }

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.brands = []
        self.total_brands_count = 0
        self.items_scraped_count = 0
        self.sku_cost_lookup = self.load_sku_cost_lookup()

    @staticmethod
    def normalize_sku(value):
        if value is None:
            return ''
        cleaned = str(value).strip().upper()
        cleaned = re.sub(r'^KAR[\s\-]+', '', cleaned)
        cleaned = re.sub(r'\s+', '', cleaned)
        return cleaned

    @staticmethod
    def compact_sku(value):
        return re.sub(r'[^A-Z0-9]', '', value or '')

    @staticmethod
    def format_number(value):
        if value is None:
            return ''
        text = str(value).strip()
        if not text:
            return ''
        try:
            num = float(text.replace(',', ''))
        except Exception:
            return text
        if num.is_integer():
            return str(int(num))
        return str(num)

    @staticmethod
    def to_float(value):
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        try:
            return float(text.replace(',', ''))
        except Exception:
            return None

    def calc_case_price(self, unit_price, case_pack):
        unit_num = self.to_float(unit_price)
        pack_num = self.to_float(case_pack)
        if unit_num is None or pack_num is None:
            return ''
        return self.format_number(unit_num * pack_num)

    def attach_ahscompany_pricing(self, item):
        ahs_price = self.format_number(item.get('price', ''))
        item['ahscompany_price'] = ahs_price
        item['ahscompany_case_price'] = self.calc_case_price(ahs_price, item.get('sku_case_pack', ''))

    @staticmethod
    def build_option_values(options):
        if not isinstance(options, dict):
            return []
        out = []
        for name, label in options.items():
            option_name = str(name).strip()
            option_label = str(label).strip()
            if not option_name or not option_label:
                continue
            out.append({
                'option_display_name': option_name,
                'label': option_label,
            })
        return out

    def resolve_price_file_path(self):
        project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        candidates = [
            os.path.join(project_root, '2026 Kartri Price List.xlsx'),
            os.path.join(project_root, '..', '..', 'ETL - process', 'ahscompany_products', '2026 Kartri Price List.xlsx'),
        ]
        for path in candidates:
            full_path = os.path.abspath(path)
            if os.path.exists(full_path):
                return full_path
        return ''

    def _get_col_index(self, headers, *aliases):
        for alias in aliases:
            if alias in headers:
                return headers[alias]
        return None

    def load_sku_cost_lookup(self):
        """
        Build lookup by SKU from 2026 Kartri Price List.xlsx.
        Keys include normalized and compact versions to absorb formatting differences.
        """
        file_path = self.resolve_price_file_path()
        if not file_path:
            self.logger.warning('2026 Kartri Price List.xlsx not found. sku_cost fields will be blank.')
            return {}

        try:
            from openpyxl import load_workbook
        except Exception as e:
            self.logger.error(f'openpyxl is unavailable. sku_cost mapping disabled: {e}')
            return {}

        lookup = {}
        try:
            wb = load_workbook(file_path, data_only=True, read_only=True)
            ws = wb.active

            header_row = None
            header_map = {}
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
                normalized = [(str(v).strip().lower() if v is not None else '') for v in row]
                if 'product id' in normalized and ('price /ea' in normalized or 'case price' in normalized):
                    header_row = row_idx
                    for idx, name in enumerate(normalized):
                        if name:
                            header_map[name] = idx
                    break

            if not header_row:
                self.logger.warning(f'Header row not found in {file_path}. sku_cost fields will be blank.')
                return {}

            product_id_idx = self._get_col_index(header_map, 'product id')
            price_each_idx = self._get_col_index(header_map, 'price /ea', 'price/ea', 'price / ea')
            case_price_idx = self._get_col_index(header_map, 'case price')
            case_pack_idx = self._get_col_index(header_map, 'case pk', 'case pack')

            if product_id_idx is None:
                self.logger.warning(f'Product ID column not found in {file_path}. sku_cost fields will be blank.')
                return {}

            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                pid_raw = row[product_id_idx] if product_id_idx < len(row) else None
                pid = str(pid_raw).strip() if pid_raw is not None else ''
                if not pid:
                    continue

                price_each = row[price_each_idx] if price_each_idx is not None and price_each_idx < len(row) else None
                case_price = row[case_price_idx] if case_price_idx is not None and case_price_idx < len(row) else None
                case_pack = row[case_pack_idx] if case_pack_idx is not None and case_pack_idx < len(row) else None

                if price_each is None and case_price is None and case_pack is None:
                    # Skip heading rows like "Replacement Liner"
                    continue

                value = {
                    'sku_cost': self.format_number(price_each),
                    'sku_case_pack': self.format_number(case_pack),
                    'sku_case_cost': self.calc_case_price(price_each, case_pack),
                }

                norm = self.normalize_sku(pid)
                compact = self.compact_sku(norm)
                if norm:
                    lookup[norm] = value
                if compact:
                    lookup[compact] = value

            self.logger.info(f'Loaded {len(lookup)} sku_cost keys from {file_path}')
            return lookup
        except Exception as e:
            self.logger.error(f'Failed to parse {file_path}: {e}')
            return {}

    def get_cost_for_sku(self, *sku_candidates):
        for sku in sku_candidates:
            norm = self.normalize_sku(sku)
            if not norm:
                continue
            compact = self.compact_sku(norm)
            if norm in self.sku_cost_lookup:
                return self.sku_cost_lookup[norm]
            if compact in self.sku_cost_lookup:
                return self.sku_cost_lookup[compact]
        return {
            'sku_cost': '',
            'sku_case_cost': '',
            'sku_case_pack': '',
        }

    def parse(self, response, **kwargs):
        """
        Entry point for crawl. Start by parsing categories on the current page.
        """
        yield from self.parse_brand_categories(response)

    def parse_brand_categories(self, response):
        """
        Parse category pages to discover child categories and product detail pages.
        If no category/product links are present, try parsing as product detail page.

        """

        # Category links: support old + current markup.
        category_links = response.css(
            '.sub-categories-format .sub-categories a[href], '
            '.subcategories_block a[href], '
            'nav.subcategories a[href]'
        )
        seen_category_urls = set()

        for category_link in category_links:
            href = category_link.css('::attr(href)').get('')
            if not href:
                continue
            if '_c_' not in href or not href.lower().endswith('.html'):
                continue

            category_url = urljoin(self.base_url, href)
            if category_url in seen_category_urls:
                continue
            seen_category_urls.add(category_url)

            next_meta = dict(response.meta)
            category_name = category_link.css('::text').get('')
            category_name = category_name.strip() if category_name else ''

            if category_name:
                if not next_meta.get('category', ''):
                    next_meta['category'] = category_name
                elif not next_meta.get('sub_category', ''):
                    next_meta['sub_category'] = category_name
                elif not next_meta.get('sub_sub_category', ''):
                    next_meta['sub_sub_category'] = category_name

            yield Request(
                url=category_url,
                headers=self.headers,
                callback=self.parse_brand_categories,
                meta=next_meta,
            )

        # Product links: support old + current listing layouts.
        product_urls = set()

        listing_products = response.css(
            '#itemsBlock .productBlockContainer .product-item, '
            '#itemsBlock .productBlockContainer'
        )
        for product in listing_products:
            href = product.css('.name a::attr(href)').get('')
            if not href:
                continue
            product_url = urljoin(self.base_url, href)
            if '_p_' in href or ('product.asp' in href and 'itemid=' in href):
                product_urls.add(product_url)

        if not product_urls:
            for href in response.css('a[href*="_p_"]::attr(href), a[href*="product.asp?itemid="]::attr(href)').getall():
                if not href:
                    continue
                product_urls.add(urljoin(self.base_url, href))

        for product_url in sorted(product_urls):
            yield Request(
                product_url,
                headers=self.headers,
                callback=self.parse_product_details,
                meta=dict(response.meta),
            )

        # If no categories and no product links, treat as direct product detail page.
        if not seen_category_urls and not product_urls:
            yield from self.parse_product_details(response)

    def parse_product_details(self, response):
        """
        Parse product detail pages to extract product information.
        """

        try:
            pid = response.css('input[name="item_id"]::attr(value)').get('')
            title = response.css('h1[itemprop="name"] ::text').get('')
            add_to_card = response.css('#Add').get('')
            description = response.css('[itemprop="description"]  ::text').getall()
            price = response.css('meta[itemprop="price"]::attr(content)').get('')
            part_number = response.css('#product_id ::text').get('')
            # Adjust the part_number by removing any prefix before a space
            raw_part_number = part_number
            part_number = part_number.split(' ', 1)[-1].strip() if ' ' in part_number else part_number
            

            if not title and not pid:
                return

            item = OrderedDict()
            item['product_id'] = pid
            item['parent_product_id'] = pid
            item['parent_title'] = title
            item['record_type'] = 'product'
            item['title'] = title
            item['variant_sku'] = part_number
            item['option_values'] = []
            item['price'] = str(price)
            item['retail_price'] = str(
                response.css('.retailprice span::text').get('').replace('$', '').replace(',', ''))
            item['brand'] = self.get_value(response, 'Brand') or ''.join(
                response.css('.breadcrumbs a ::text').getall()[1:2])
            item['stock_status'] = 'In Stock' if add_to_card else 'Out Of Stock'
            item['part_number'] = part_number
            item.update(self.get_cost_for_sku(part_number, raw_part_number))
            self.attach_ahscompany_pricing(item)
            item['images'] = self.get_product_images(response)
            item['description'] = '\n'.join(
                [item.strip() for item in description if item.strip()]) if description else ''
            item['category'] = ''.join(response.css('.breadcrumbs a ::text').getall()[2:3])
            item['sub_category'] = ', '.join(response.css('.breadcrumbs a ::text').getall()[3:])
            item['size'] = self.get_value(response, 'Size')
            item['type'] = self.get_value(response, 'Type')
            item['material'] = self.get_value(response, 'Material')
            item['options'] = {}
            item['url'] = response.css('link[rel="canonical"] ::attr(href)').get('') or response.url

            variants = response.css('#divOptionsBlock option')

            if not variants:
                self.items_scraped_count += 1
                print(f'Current items scraped: {self.items_scraped_count}')

                yield item

                return

            # Emit one parent product row, then variant rows keyed to this product_id.
            self.items_scraped_count += 1
            print(f'Current items scraped: {self.items_scraped_count}')
            yield copy.deepcopy(item)

            # Parse Variants
            # if there is one single type of variant

            if len(response.css('#divOptionsBlock .dropdown-format').getall()) == 1:
                variants = response.css('#divOptionsBlock option')[1:]
                label = response.css('#divOptionsBlock option::text').get('').strip()

                for variant in variants:
                    variant_item = copy.deepcopy(item)
                    options = self.get_variant_price(response, variant, price, part_number, label)
                    variant_item['price'] = options.get('price', '')
                    p_number = options.get('part_number', '')
                    p_number = p_number.split(' ', 1)[-1].strip() if ' ' in p_number else p_number
                    variant_item['part_number'] = p_number
                    variant_item['record_type'] = 'variant'
                    variant_item['parent_product_id'] = pid
                    variant_item['parent_title'] = title
                    variant_item['variant_sku'] = p_number
                    variant_item.update(self.get_cost_for_sku(p_number, options.get('part_number', ''), raw_part_number))
                    variant_item['options'] = options.get('variant_name', {})
                    variant_item['option_values'] = self.build_option_values(variant_item['options'])
                    self.attach_ahscompany_pricing(variant_item)
                    self.items_scraped_count += 1
                    print(f'Current items scraped: {self.items_scraped_count}')

                    yield variant_item
            else:
                # If there are 2 or 3 types of variants in dropdowns selection
                first_variant_group = response.css('#divOptionsBlock .container > :nth-child(1) option')[1:]
                first_variant_label = response.css('#divOptionsBlock .container > :nth-child(1) option::text').get(
                    '').strip()
                second_variant_group = response.css('#divOptionsBlock .container > :nth-child(3) option')[1:]
                second_variant_label = response.css('#divOptionsBlock .container > :nth-child(3) option::text').get(
                    '').strip()
                third_variant_group = response.css('#divOptionsBlock .container > :nth-child(5) option')[1:]
                third_variant_label = response.css('#divOptionsBlock .container > :nth-child(5) option::text').get(
                    '').strip()

                for first_variant in first_variant_group:
                    for second_variant in second_variant_group:
                        for third_variant in third_variant_group:
                            variant_item = copy.deepcopy(item)
                            options = self.get_variant_group(response, first_variant, first_variant_label,
                                                             second_variant, second_variant_label, price,
                                                             third_variant, third_variant_label, part_number)

                            variant_item['price'] = options.get('price', '')
                            p_number = options.get('part_number', '')
                            p_number = p_number.split(' ', 1)[-1].strip() if ' ' in p_number else p_number
                            variant_item['part_number'] = p_number
                            variant_item['record_type'] = 'variant'
                            variant_item['parent_product_id'] = pid
                            variant_item['parent_title'] = title
                            variant_item['variant_sku'] = p_number
                            variant_item.update(self.get_cost_for_sku(
                                variant_item.get('part_number', ''),
                                raw_part_number,
                            ))
                            variant_item['options'] = options.get('variant_name', {})
                            variant_item['option_values'] = self.build_option_values(variant_item['options'])
                            self.attach_ahscompany_pricing(variant_item)
                            self.items_scraped_count += 1
                            print(f'Current items scraped: {self.items_scraped_count}')

                            yield variant_item

        except Exception as e:
            print(f'Error parsing product detail: {response.url}')
            self.logger.error(f'Error parsing product detail: {response.url} and Error : {e}')
            return

    def get_value(self, response, key):
        elements = response.css(f'.breadcrumbsBlock ul li:contains("{key}") ::text').getall()
        if elements:
            return elements[-1]
        else:
            return ''

    def get_product_images(self, response):
        images = response.css('.slides .prod-thumb a::attr(href)').getall() or []
        images = images or response.css('.main-image a::attr(href)').getall() or []
        return images

    def get_variant_price(self, response, variant, price, part_number, label):
        variant_value = variant.css('::attr(value)').get('')
        variant_name = variant.css('::text').get('').strip()

        if not variant_value:
            info = {
                'price': str(price),
                'part_number': part_number,
                'variant_name': {label: variant_name}
            }

            return info

        price_value = response.css(f'.container input[name="price_{variant_value}"]::attr(value)').get('')
        if price_value == 0 and price_value is None:
            variant_price = price
        else:
            variant_price = float(price) + float(price_value)

        script_tag = response.css(f'script[type="text/javascript"]:contains("{variant_value}")').get('')
        if script_tag:
            p_number = script_tag.split(variant_value)[1].split('=')[1].split(';')[0].strip().replace("'",
                                                                                                      "") or part_number
        else:
            p_number = part_number
        info = {
            'price': str(variant_price),
            'part_number': p_number,
            'variant_name': {label: variant_name}
        }

        return info

    def get_variant_group(self, response, first_variant, first_variant_label, second_variant, second_variant_label,
                          price, third_variant, third_variant_label, part_number):

        f_variant = self.get_variant_price(response, first_variant, price, part_number, first_variant_label)
        sec_variant = self.get_variant_price(response, second_variant, price, part_number, second_variant_label)
        thi_variant = self.get_variant_price(response, third_variant, price, part_number, third_variant_label)
        p_number = f_variant.get('part_number', '')
        f_variant_price = f_variant.get('price', '')

        info = {
            'part_number': p_number,
            'price': str(f_variant_price),
            'variant_name': {
                **f_variant.get('variant_name', {}),
                **sec_variant.get('variant_name', {}),
                **thi_variant.get('variant_name', {})}
        }

        return info

    @classmethod
    def from_crawler(cls, crawler, *args, **kwargs):
        spider = super(AshSpider, cls).from_crawler(crawler, *args, **kwargs)
        crawler.signals.connect(spider.spider_idle, signal=signals.spider_idle)
        return spider

    def spider_idle(self):
        """
        Handle spider idle state by crawling next brand if available.
        """

        print(f'\n\nTotal {self.items_scraped_count} items scraped')
        print(f'\n\n{len(self.brands)}/{self.total_brands_count} Brands left to Scrape\n\n')

        if self.brands:
            brand = self.brands.pop(0)
            brand_name = brand.get('name', '')
            brand_url = brand.get('url', '')

            req = Request(url=brand_url,
                          callback=self.parse_brand_categories, dont_filter=True,
                          meta={'handle_httpstatus_all': True, 'brand': brand_name})

            try:
                self.crawler.engine.crawl(req)  # For latest Python version
            except TypeError:
                self.crawler.engine.crawl(req, self)  # For old Python version < 10
