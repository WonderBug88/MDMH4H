import os
import base64
import psycopg2
import pandas as pd
from dateutil import parser
from dotenv import load_dotenv
from openpyxl import load_workbook
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow

BASE_DIR = os.path.dirname(os.path.dirname(
    os.path.dirname(os.path.abspath(__file__))))

# Load environment variables
load_dotenv(dotenv_path=os.path.join(BASE_DIR, '.env'))

# Example of accessing environment variables
db_config = {
    'dbname': os.getenv('DB_NAME'),
    'user': os.getenv('DB_USER'),
    'password': os.getenv('DB_PASSWORD'),
    'host': os.getenv('DB_HOST'),
    'port': os.getenv('DB_PORT'),
}

# Gmail API Scopes and Files from .env
SCOPES = ["https://www.googleapis.com/auth/gmail.readonly"]
CREDENTIALS_FILE = os.getenv('GMAIL_CREDENTIALS_FILE')
TOKEN_FILE = os.getenv('GMAIL_TOKEN_FILE')

print("SCOPES", SCOPES)
print("CREDENTIALS_FILE", CREDENTIALS_FILE)
print("TOKEN_FILE", TOKEN_FILE)

def authenticate_gmail():
    """Authenticate and return the Gmail API service."""
    creds = None
    if os.path.exists(TOKEN_FILE):
        creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
    if not creds or not creds.valid:
        flow = InstalledAppFlow.from_client_secrets_file(
            CREDENTIALS_FILE, SCOPES)
        creds = flow.run_local_server(port=0)
        with open(TOKEN_FILE, 'w') as token:
            token.write(creds.to_json())
    return build('gmail', 'v1', credentials=creds)


def get_all_email_attachments(service, user_id, query):
    """Download all attachments from emails matching the query."""
    try:
        results = service.users().messages().list(userId=user_id, q=query).execute()
        messages = results.get('messages', [])
        if not messages:
            print("No messages found.")
            return []

        downloaded_files = []
        for message in messages:
            msg = service.users().messages().get(
                userId=user_id, id=message['id']).execute()
            payload = msg.get('payload', {})
            headers = payload.get('headers', [])

            # Extract the email date
            email_date = next(
                (header['value'] for header in headers if header['name'] == 'Date'), None)
            try:
                email_date = parser.parse(email_date)
                formatted_date = email_date.strftime('%Y-%m-%d')
            except Exception as e:
                print(f"Failed to parse date: {email_date}. Error: {e}")
                formatted_date = 'unknown_date'

            # Process attachments
            parts = payload.get('parts', [])
            for part in parts:
                if part.get('filename') and part.get('mimeType') == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
                    attachment_id = part['body'].get('attachmentId')
                    attachment = service.users().messages().attachments().get(
                        userId=user_id, messageId=message['id'], id=attachment_id
                    ).execute()
                    data = base64.urlsafe_b64decode(attachment.get('data'))

                    # Save the attachment
                    filename = f"{formatted_date}_{part['filename']}"
                    file_path = os.path.join('downloads', filename)
                    with open(file_path, 'wb') as f:
                        f.write(data)
                        print(f'Attachment saved: {file_path}')

                    # Append file path and timestamp
                    downloaded_files.append((file_path, formatted_date))
        return downloaded_files
    except HttpError as error:
        print(f'An error occurred: {error}')
        return []


def process_and_upload(file_path, email_date):
    """Process the downloaded file and upload it to the database."""
    try:
        # Unmerge the Excel file
        wb = load_workbook(file_path)
        ws = wb.active
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_cell in merged_ranges:
            ws.unmerge_cells(str(merged_cell))
        temp_file = file_path.replace(".xlsx", "_unmerged.xlsx")
        wb.save(temp_file)

        # Process the data with pandas
        df = pd.read_excel(
            temp_file, header=None).iloc[5:].reset_index(drop=True)
        df = df[[1, 3, 4, 9, 12]]
        df.columns = ['Parent_Product', 'SKU', 'Item_Type', 'Inventory', 'UOM']
        df['Parent_Product'] = df['Parent_Product'].ffill()
        words_to_remove = ['UOM', 'Available', 'ItemName', 'ItemCode']
        df.replace(to_replace=words_to_remove,
                   value='', regex=True, inplace=True)

        # Remove rows where all key columns are empty
        df = df[
            (df['SKU'].notna() & (df['SKU'] != '')) |
            (df['Item_Type'].notna() & (df['Item_Type'] != '')) |
            (df['Inventory'].notna() & (df['Inventory'] != '')) |
            (df['UOM'].notna() & (df['UOM'] != ''))
        ]
        df['UOM'] = df['UOM'].str.strip().str.replace(
            r'\s+', ' ', regex=True).str.upper()
        uom_mapping = {
            'DOZ': 'Dozens', 'DOZEN': 'Dozens', 'EA': 'Each', 'CASE': 'Case',
            'YDS': 'Yards', 'EACH': 'Each', 'YARD': 'Yards'
        }
        df['UOM'] = df['UOM'].map(uom_mapping).fillna('Unknown')

        # Use the email's date as the timestamp
        df['Timestamp'] = email_date

        print("Rows after filtering:")
        print(df.head())
        print(f"Remaining rows: {len(df)}")

        # Upload to PostgreSQL
        conn = psycopg2.connect(**db_config)
        cursor = conn.cursor()
        for _, row in df.iterrows():
            cursor.execute(
                """
                INSERT INTO ganesh.inventory (Parent_Product, SKU, Item_Type, Inventory, UOM, Timestamp)
                VALUES (%s, %s, %s, %s, %s, %s);
                """,
                (row['Parent_Product'], row['SKU'], row['Item_Type'],
                 row['Inventory'], row['UOM'], row['Timestamp'])
            )
        conn.commit()
        print("Data inserted successfully!")
    except Exception as e:
        print(f"Error: {e}")
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'conn' in locals() and conn:
            conn.close()
        print("Script execution complete.")


def start_all_process():
    os.makedirs('downloads', exist_ok=True)
    service = authenticate_gmail()
    if service:
        query = 'from:vchawhan@ganeshmills.com subject:"RE: Stock Status Report" has:attachment'
        downloaded_files = get_all_email_attachments(service, 'me', query)
        print(f"Downloaded {len(downloaded_files)} attachments.")
        for file_path, email_date in downloaded_files:
            process_and_upload(file_path, email_date)

if __name__ == '__main__':
    start_all_process()